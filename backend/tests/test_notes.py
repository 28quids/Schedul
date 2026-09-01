"""Notes layer from the practice down to one document, and say where they came from.

The order is the point. A reader expects the standing compliance paragraph
before "radiant panels are to be sized with a 55degC flow", and a schedule that
has to say something different has to be able to without editing the practice's
wording for everybody else.
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

from schedul.core.notes import resolve_notes, seed_from

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")


class TestResolution:
    def test_layers_print_general_to_specific_and_are_numbered_once(self):
        resolved = resolve_notes(
            organisation=["house one", "house two"],
            project=["job one"],
            type_notes=["panels"],
        )
        assert [n.text for n in resolved] == ["house one", "house two", "job one", "panels"]
        assert [n.number for n in resolved] == [1, 2, 3, 4]
        assert [n.source for n in resolved] == [
            "organisation", "organisation", "project", "type",
        ]

    def test_a_schedule_with_its_own_notes_replaces_the_layers(self):
        resolved = resolve_notes(
            organisation=["house"], project=["job"], type_notes=["panels"],
            schedule=["only this"],
        )
        assert [n.text for n in resolved] == ["only this"]
        assert [n.source for n in resolved] == ["schedule"]

    def test_deliberately_no_notes_is_not_the_same_as_inheriting(self):
        assert resolve_notes(organisation=["house"], schedule=[]) == []
        assert len(resolve_notes(organisation=["house"], schedule=None)) == 1

    def test_blank_lines_never_reach_a_document(self):
        resolved = resolve_notes(organisation=["a", "", "   ", "b"])
        assert [n.text for n in resolved] == ["a", "b"]

    def test_diverging_starts_from_what_it_says_now(self):
        resolved = resolve_notes(organisation=["house"], type_notes=["panels"])
        assert seed_from(resolved) == ["house", "panels"]


@pytest.fixture()
def client(tmp_path, monkeypatch) -> TestClient:
    monkeypatch.setenv("SCHEDUL_DATABASE_URL", f"sqlite:///{tmp_path / 'notes.db'}")
    import schedul.db.session as session_module

    session_module.SessionLocal = None
    session_module.init_db(f"sqlite:///{tmp_path / 'notes.db'}")
    from schedul.api.main import app

    return TestClient(app)


@pytest.fixture()
def project(client) -> dict:
    return client.post("/api/projects", json={"number": "CM1", "name": "J"}).json()


@pytest.fixture()
def schedule(client, project) -> str:
    building = project["buildings"][0]["id"]
    result = client.post(
        f"/api/projects/{project['id']}/buildings/{building}/schedules",
        json={"code": "RADPANEL"},
    ).json()
    return result["buildings"][0]["schedules"][0]["id"]


class TestThroughTheApi:
    def test_a_schedule_shows_every_layer_and_where_each_came_from(self, client, schedule):
        notes = client.get(f"/api/schedules/{schedule}/notes").json()
        sources = {n["source"] for n in notes["note_layers"]}
        assert sources == {"organisation", "type"}
        assert notes["notes_customised"] is False

    def test_project_notes_sit_between_the_practice_and_the_type(
        self, client, project, schedule
    ):
        client.put(f"/api/projects/{project['id']}", json={
            "number": "CM1", "name": "J", "notes": ["site rule"],
        })
        notes = client.get(f"/api/schedules/{schedule}/notes").json()
        texts = [n["text"] for n in notes["note_layers"]]
        sources = [n["source"] for n in notes["note_layers"]]

        assert "site rule" in texts
        assert sources.index("project") > sources.index("organisation")
        assert sources.index("project") < sources.index("type")

    def test_saving_a_project_without_notes_does_not_blank_them(self, client, project):
        client.put(f"/api/projects/{project['id']}", json={
            "number": "CM1", "name": "J", "notes": ["site rule"],
        })
        after = client.put(f"/api/projects/{project['id']}", json={
            "number": "CM1", "name": "Renamed",
        }).json()
        assert after["notes"] == ["site rule"]

    def test_a_schedule_can_diverge_and_come_back(self, client, schedule):
        inherited = client.get(f"/api/schedules/{schedule}/notes").json()

        taken = client.post(f"/api/schedules/{schedule}/notes/customise").json()
        assert taken["notes_customised"] is True
        assert taken["notes"] == inherited["notes"], "diverging starts from what it said"

        edited = client.put(f"/api/schedules/{schedule}/notes", json={
            "notes": ["this one is different"],
        }).json()
        assert edited["notes"] == ["this one is different"]
        assert [n["source"] for n in edited["note_layers"]] == ["schedule"]

        # The layers are still reported, so the way back is visible.
        assert edited["layers"]["organisation"]
        assert [n["text"] for n in edited["inherited"]] == inherited["notes"]

        reverted = client.put(f"/api/schedules/{schedule}/notes", json={"notes": None}).json()
        assert reverted["notes_customised"] is False
        assert reverted["notes"] == inherited["notes"]

    def test_one_schedule_diverging_leaves_its_neighbours_alone(
        self, client, project, schedule
    ):
        building = project["buildings"][0]["id"]
        other = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()["buildings"][0]["schedules"]
        other_id = next(s["id"] for s in other if s["code"] == "MVHR")

        client.put(f"/api/schedules/{schedule}/notes", json={"notes": ["mine"]})
        assert client.get(f"/api/schedules/{other_id}/notes").json()["notes_customised"] is False

    def test_the_grid_carries_the_resolved_notes(self, client, schedule):
        grid = client.get(f"/api/schedules/{schedule}").json()
        assert grid["notes"], "the editor shows what will print"
        assert grid["note_layers"][0]["source"] == "organisation"
        assert grid["notes_customised"] is False

    def test_house_notes_reach_every_schedule_that_has_not_diverged(self, client, schedule):
        client.put("/api/settings", json={"general_notes": ["one house note"]})
        notes = client.get(f"/api/schedules/{schedule}/notes").json()
        assert notes["notes"][0] == "one house note"

        client.put(f"/api/schedules/{schedule}/notes", json={"notes": ["mine"]})
        client.put("/api/settings", json={"general_notes": ["changed again"]})
        assert client.get(f"/api/schedules/{schedule}/notes").json()["notes"] == ["mine"], (
            "a schedule that has taken its notes over is not reached"
        )


class TestNotesReachTheDocuments:
    def test_the_workbook_prints_the_resolved_notes(self, client, project, schedule):
        from openpyxl import load_workbook
        import io

        client.put(f"/api/projects/{project['id']}", json={
            "number": "CM1", "name": "J", "notes": ["a note only this job has"],
        })
        response = client.get(f"/api/schedules/{schedule}/export.xlsx")
        assert response.status_code == 200
        block = load_workbook(io.BytesIO(response.content))["Schedule"]["A2"].value

        assert "a note only this job has" in block
        assert block.index("read in conjunction") < block.index("a note only this job has")

    def test_a_schedules_own_notes_are_what_it_exports(self, client, schedule):
        from openpyxl import load_workbook
        import io

        client.put(f"/api/schedules/{schedule}/notes", json={"notes": ["just this one"]})
        response = client.get(f"/api/schedules/{schedule}/export.xlsx")
        block = load_workbook(io.BytesIO(response.content))["Schedule"]["A2"].value

        assert block == "General Notes:\n[1] just this one"

    def test_issuing_freezes_the_notes_as_they_read_then(self, client, schedule):
        revision = client.post(f"/api/schedules/{schedule}/revisions", json={
            "code": "P01", "status": "S2 - Suitable for Information",
        }).json()[0]
        client.post(f"/api/schedules/{schedule}/revisions/{revision['id']}/issue")

        client.put("/api/settings", json={"general_notes": ["reworded afterwards"]})
        snapshot = client.get(
            f"/api/schedules/{schedule}/revisions/{revision['id']}/snapshot"
        ).json()

        assert "reworded afterwards" not in snapshot["notes"]
        assert any("read in conjunction" in n for n in snapshot["notes"])


class TestGettingTheHouseNotesBack:
    """A practice that has emptied its general notes has somewhere to go.

    Storing an empty list is a real answer -- a practice may print no notes --
    so the defaults deliberately do not creep back on their own. That makes the
    way back an offer the screen has to be able to make, which means the API has
    to hand out the built-in wording rather than the screen inventing it.
    """

    def test_a_fresh_practice_starts_with_the_built_in_notes(self, client):
        settings = client.get("/api/settings").json()
        assert settings["house_standard"]["general_notes"], (
            "the notes a schedule prints must be editable where the screen says they are"
        )

    def test_the_built_in_wording_is_offered_even_once_the_notes_are_emptied(self, client):
        from schedul.core.house import DEFAULT_GENERAL_NOTES

        client.put("/api/settings", json={"general_notes": []})
        settings = client.get("/api/settings").json()
        assert settings["house_standard"]["general_notes"] == []
        assert settings["default_general_notes"] == list(DEFAULT_GENERAL_NOTES), (
            "emptying the notes must not be a one-way door"
        )

    def test_restoring_them_puts_them_back_on_a_schedule(self, client, schedule):
        from schedul.core.house import DEFAULT_GENERAL_NOTES

        client.put("/api/settings", json={"general_notes": []})
        assert not [
            n for n in client.get(f"/api/schedules/{schedule}/notes").json()["note_layers"]
            if n["source"] == "organisation"
        ]

        client.put("/api/settings", json={"general_notes": list(DEFAULT_GENERAL_NOTES)})
        restored = [
            n["text"]
            for n in client.get(f"/api/schedules/{schedule}/notes").json()["note_layers"]
            if n["source"] == "organisation"
        ]
        assert restored == list(DEFAULT_GENERAL_NOTES)
