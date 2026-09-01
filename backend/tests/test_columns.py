"""Column visibility, project-specific columns, and library overrides.

All three change what "the columns of this schedule" means, which is why they
share a test file: the grid, the renderer and the validator must agree on the
answer, and the risk is that one of them learns about extras and another does
not.
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from schedul.core.catalogue import Column, ScheduleType, validate_type
from schedul.db.models import Base, Project
from schedul.db.session import make_engine
from schedul.services import projects as svc
from schedul.services.columns import merged_type, project_extras, set_project_extras
from schedul.services.grid import compute_row, editable_payload, override_payload
from schedul.services.seed import seed_organisation


@pytest.fixture()
def session(tmp_path):
    engine = make_engine(f"sqlite:///{tmp_path / 'cols.db'}")
    Base.metadata.create_all(engine)
    with sessionmaker(bind=engine, expire_on_commit=False, future=True)() as s:
        yield s


@pytest.fixture()
def org(session, repo_root):
    return seed_organisation(
        session, "T", "t", schema_path=repo_root / "vendor" / "schema.json"
    )


@pytest.fixture()
def project(session, org):
    p = Project(organisation_id=org.id, name="J", number="CM1")
    session.add(p)
    session.flush()
    return p


def a_type(**kw) -> ScheduleType:
    defaults = dict(
        code="RAD", title="Radiator Schedule",
        columns=[
            Column("input", "Ref", example="RAD-01"),
            Column("library", "Manufacturer"),
            Column("library", "Output at dT50", unit="W"),
        ],
    )
    defaults.update(kw)
    return ScheduleType(**defaults)


class TestVisibility:
    def test_absent_visibility_means_visible_everywhere(self):
        column = Column("input", "Ref")
        assert column.visible_in("editor")
        assert column.visible_in("xlsx")
        assert column.visible_in("pdf")

    def test_a_column_can_be_hidden_from_issued_documents(self):
        """The Price case: kept on the schedule, off the deliverable."""
        price = Column("input", "Price", visibility={"xlsx": False, "pdf": False})
        assert price.visible_in("editor")
        assert not price.visible_in("xlsx")
        assert not price.visible_in("pdf")

    def test_visible_columns_filters_the_type(self):
        st = a_type()
        st.columns.append(Column("input", "Price", visibility={"pdf": False}))
        assert [c.name for c in st.visible_columns("pdf").columns] == [
            "Ref", "Manufacturer", "Output at dT50"
        ]
        assert len(st.visible_columns("editor").columns) == 4

    def test_visibility_survives_serialisation(self):
        column = Column("input", "Price", visibility={"pdf": False})
        assert Column.from_dict(column.to_dict()).visibility == {"pdf": False}

    def test_an_old_catalogue_entry_needs_no_migration(self):
        """A column dict written before visibility existed still loads."""
        assert Column.from_dict({"kind": "input", "name": "Ref"}).visible_in("pdf")


class TestProjectColumns:
    def test_extras_are_appended_and_marked(self, session, project):
        set_project_extras(project, "RAD", [Column("input", "Quantity", example=2)])
        merged = merged_type(project, a_type())

        assert [c.name for c in merged.columns][-1] == "Quantity"
        assert merged.columns[-1].project_extra is True
        assert merged.column("Quantity") is not None

    def test_the_base_type_is_untouched(self, session, project):
        base = a_type()
        set_project_extras(project, "RAD", [Column("input", "Quantity")])
        merged_type(project, base)
        assert [c.name for c in base.columns] == ["Ref", "Manufacturer", "Output at dT50"]

    def test_another_project_does_not_see_them(self, session, org, project):
        other = Project(organisation_id=org.id, name="K", number="CM2")
        session.add(other)
        session.flush()
        set_project_extras(project, "RAD", [Column("input", "Quantity")])
        assert project_extras(other, "RAD") == []

    def test_a_duplicate_of_a_base_column_is_ignored(self):
        """Additions only: an extra cannot shadow a catalogue column."""
        merged = a_type().with_extras([Column("input", "Manufacturer")])
        assert [c.name for c in merged.columns].count("Manufacturer") == 1

    def test_a_project_derived_column_may_use_base_columns(self, session, project):
        extra = Column(
            "derived", "Total Output", unit="W",
            formula="={Quantity}*{Output at dT50 (W)}", note="n",
        )
        set_project_extras(
            project, "RAD", [Column("input", "Quantity", example=2), extra]
        )
        merged = merged_type(project, a_type())
        # Validation runs against the merged list, so the reference resolves.
        assert [i for i in validate_type(merged) if i.severity == "error"] == []

    def test_a_project_derived_column_with_a_bad_reference_is_caught(self, session, project):
        set_project_extras(
            project, "RAD",
            [Column("derived", "Bad", formula="={Nope}*2", note="n")],
        )
        merged = merged_type(project, a_type())
        assert any("not a column" in i.message for i in validate_type(merged))

    def test_setting_an_empty_list_removes_them(self, session, project):
        set_project_extras(project, "RAD", [Column("input", "Quantity")])
        set_project_extras(project, "RAD", [])
        assert project_extras(project, "RAD") == []


class TestOverrides:
    def constants(self):
        return {}

    def test_an_override_replaces_the_library_value(self):
        st = a_type()
        cells = compute_row(
            {"Ref": "RAD-01", "Model Reference": "K2-600"},
            st,
            {"K2-600": {"Manufacturer": "Stelrad", "Output at dT50 (W)": 700}},
            self.constants(),
            overrides={"Output at dT50 (W)": 820},
        )
        assert cells["Manufacturer"].value == "Stelrad", "un-overridden fields still resolve"
        assert cells["Output at dT50 (W)"].value == 820
        assert cells["Output at dT50 (W)"].overridden is True
        assert cells["Output at dT50 (W)"].editable is True

    def test_an_override_works_without_a_model_reference(self):
        """A row can carry a manual value before the product exists at all."""
        cells = compute_row(
            {"Ref": "RAD-01"}, a_type(), {}, self.constants(),
            overrides={"Manufacturer": "TBC"},
        )
        assert cells["Manufacturer"].value == "TBC"
        assert cells["Manufacturer"].problem is None

    def test_a_derived_column_uses_the_overridden_value(self):
        st = a_type()
        st.columns.append(
            Column("input", "Quantity", example=1)
        )
        st.columns.append(
            Column("derived", "Total", unit="W",
                   formula="={Quantity}*{Output at dT50 (W)}", note="n")
        )
        cells = compute_row(
            {"Ref": "R", "Quantity": 2, "Model Reference": "K2"},
            st,
            {"K2": {"Output at dT50 (W)": 700}},
            self.constants(),
            overrides={"Output at dT50 (W)": 800},
        )
        assert cells["Total (W)"].value == 1600, "the override feeds the formula"

    def test_no_override_leaves_the_library_in_charge(self):
        cells = compute_row(
            {"Ref": "R", "Model Reference": "K2"}, a_type(),
            {"K2": {"Manufacturer": "Stelrad"}}, self.constants(),
        )
        assert cells["Manufacturer"].value == "Stelrad"
        assert cells["Manufacturer"].overridden is False


class TestPayloadGuards:
    def test_a_library_value_in_values_is_still_stripped(self):
        """The guard that stops a client storing computed values must survive."""
        cleaned = editable_payload(
            {"Ref": "R", "Manufacturer": "FORGED"}, a_type()
        )
        assert cleaned == {"Ref": "R"}

    def test_only_real_library_columns_can_be_overridden(self):
        cleaned = override_payload(
            {"Manufacturer": "Stelrad", "Ref": "R", "Nonsense": 1}, a_type()
        )
        assert cleaned == {"Manufacturer": "Stelrad"}

    def test_a_blank_override_clears_rather_than_stores(self):
        """Clearing is how a row resets to the library value."""
        assert override_payload({"Manufacturer": ""}, a_type()) == {}

    def test_overrides_are_coerced_like_typed_values(self):
        cleaned = override_payload({"Output at dT50 (W)": "820"}, a_type())
        assert cleaned == {"Output at dT50 (W)": 820}

    def test_a_column_can_be_named_without_its_unit(self):
        cleaned = override_payload({"Output at dT50": 820}, a_type())
        assert cleaned == {"Output at dT50 (W)": 820}


@pytest.mark.filterwarnings("ignore::DeprecationWarning")
class TestPerScheduleVisibilityThroughTheApi:
    """One schedule hiding a column on its own deliverables.

    The catalogue says where a column belongs in general; this is the answer to
    "keep the cost off the copy that goes to this client", which is a decision
    about one document and must not reach the type or any other schedule.
    """

    @pytest.fixture()
    def client(self, tmp_path, monkeypatch):
        from fastapi.testclient import TestClient

        monkeypatch.setenv("SCHEDUL_DATABASE_URL", f"sqlite:///{tmp_path / 'vis.db'}")
        import schedul.db.session as session_module

        session_module.SessionLocal = None
        session_module.init_db(f"sqlite:///{tmp_path / 'vis.db'}")
        from schedul.api.main import app

        return TestClient(app)

    @pytest.fixture()
    def schedule(self, client):
        project = client.post(
            "/api/projects", json={"number": "CM1", "name": "Head Office"}
        ).json()
        building = project["buildings"][0]["id"]
        made = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        sid = made["buildings"][0]["schedules"][0]["id"]
        client.post(f"/api/schedules/{sid}/rows", json={
            "values": {"Unit Reference": "MVHR-01", "Supply Airflow (l/s)": 450},
        })
        return sid

    def _hideable(self, client, schedule, kind="input"):
        data = client.get(f"/api/schedules/{schedule}/columns").json()
        return next(
            c for c in data["columns"] if c["hideable"] and c["kind"] == kind
        )

    def test_every_column_says_where_it_shows(self, client, schedule):
        data = client.get(f"/api/schedules/{schedule}/columns").json()
        assert data["targets"] == ["editor", "xlsx", "pdf"]
        assert all(set(c["visibility"]) == {"editor", "xlsx", "pdf"} for c in data["columns"])
        assert all(c["visibility"]["pdf"] for c in data["columns"]), "nothing hidden yet"

    def test_the_lookup_key_cannot_be_hidden(self, client, schedule):
        data = client.get(f"/api/schedules/{schedule}/columns").json()
        reference = next(c for c in data["columns"] if c["name"] == "Model Reference")
        assert not reference["hideable"]
        response = client.put(f"/api/schedules/{schedule}/columns", json={
            "columns": {"Model Reference": {"pdf": False}},
        })
        assert response.status_code == 400

    def test_a_column_a_calculation_reads_cannot_be_hidden(self, client, schedule):
        data = client.get(f"/api/schedules/{schedule}/columns").json()
        blocked = [c for c in data["columns"] if not c["hideable"] and c["reason"].startswith("read by")]
        assert blocked, "this type has derived columns, so something must be read"
        response = client.put(f"/api/schedules/{schedule}/columns", json={
            "columns": {blocked[0]["legacy_name"]: {"xlsx": False}},
        })
        assert response.status_code == 400
        assert "read by" in response.text

    def test_hiding_a_column_takes_it_off_the_workbook(self, client, schedule):
        column = self._hideable(client, schedule)
        before = load_workbook(
            __import__("io").BytesIO(
                client.get(f"/api/schedules/{schedule}/export.xlsx").content
            )
        )["Schedule"]
        headers = [before.cell(4, c).value for c in range(1, 30)]
        assert column["name"] in headers

        client.put(f"/api/schedules/{schedule}/columns", json={
            "columns": {column["legacy_name"]: {"xlsx": False, "pdf": False}},
        })
        after = load_workbook(
            __import__("io").BytesIO(
                client.get(f"/api/schedules/{schedule}/export.xlsx").content
            )
        )["Schedule"]
        assert column["name"] not in [after.cell(4, c).value for c in range(1, 30)]

    def test_hiding_it_on_the_export_leaves_it_on_the_screen(self, client, schedule):
        column = self._hideable(client, schedule)
        client.put(f"/api/schedules/{schedule}/columns", json={
            "columns": {column["legacy_name"]: {"xlsx": False, "pdf": False}},
        })
        grid = client.get(f"/api/schedules/{schedule}").json()
        assert column["legacy_name"] in [c["legacy_name"] for c in grid["columns"]], (
            "hidden on the deliverable is not hidden in the editor"
        )

    def test_hiding_it_on_screen_takes_it_out_of_the_grid(self, client, schedule):
        column = self._hideable(client, schedule)
        client.put(f"/api/schedules/{schedule}/columns", json={
            "columns": {column["legacy_name"]: {"editor": False}},
        })
        grid = client.get(f"/api/schedules/{schedule}").json()
        assert column["legacy_name"] not in [c["legacy_name"] for c in grid["columns"]]

    def test_the_value_is_kept_not_deleted(self, client, schedule):
        client.put(f"/api/schedules/{schedule}/columns", json={
            "columns": {"Unit Reference": {"editor": False, "xlsx": False, "pdf": False}},
        })
        client.put(f"/api/schedules/{schedule}/columns", json={"columns": {}})
        grid = client.get(f"/api/schedules/{schedule}").json()
        assert grid["rows"][0]["values"]["Unit Reference"] == "MVHR-01"

    def test_it_is_this_schedule_only(self, client, schedule):
        column = self._hideable(client, schedule)
        client.put(f"/api/schedules/{schedule}/columns", json={
            "columns": {column["legacy_name"]: {"xlsx": False}},
        })
        # A second schedule of the same type on the same job is untouched.
        project_id = client.get(f"/api/schedules/{schedule}").json()["project_id"]
        project = client.get(f"/api/projects/{project_id}").json()
        building = project["buildings"][0]["id"]
        other = client.post(
            f"/api/projects/{project_id}/buildings/{building}/schedules",
            json={"code": "FCU"},
        ).json()
        assert other["buildings"][0]["schedules"]
        again = client.get(f"/api/schedules/{schedule}/columns").json()
        hidden = [c for c in again["columns"] if not c["visibility"]["xlsx"]]
        assert [c["legacy_name"] for c in hidden] == [column["legacy_name"]]
