"""Issuing a revision freezes it, and nothing later can change what it said.

This is the liability test. A schedule that has gone to a contractor must keep
saying what it said, even after somebody corrects the equipment library or fixes
a formula in the designer.
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")


@pytest.fixture()
def client(tmp_path, monkeypatch) -> TestClient:
    monkeypatch.setenv("SCHEDUL_DATABASE_URL", f"sqlite:///{tmp_path / 'issue.db'}")
    import schedul.db.session as session_module

    session_module.SessionLocal = None
    session_module.init_db(f"sqlite:///{tmp_path / 'issue.db'}")
    from schedul.api.main import app

    return TestClient(app)


@pytest.fixture()
def schedule(client) -> str:
    project = client.post("/api/projects", json={
        "name": "J", "number": "CM1", "client": "C",
    }).json()
    building = project["buildings"][0]["id"]
    result = client.post(
        f"/api/projects/{project['id']}/buildings/{building}/schedules",
        json={"code": "MVHR"},
    ).json()
    sid = result["buildings"][0]["schedules"][0]["id"]

    client.post("/api/library", json={
        "type_code": "MVHR", "model_reference": "SYS-1",
        "values": {"Manufacturer": "Systemair", "Length (mm)": 1200},
    })
    client.post(f"/api/schedules/{sid}/rows", json={"values": {
        "Unit Reference": "MVHR-01", "Model Reference": "SYS-1",
        "Supply Airflow (l/s)": 450, "Extract Airflow (l/s)": 450,
        "Total Power Input (W)": 396,
    }})
    return sid


def add_revision(client, sid, code="P01", **kw):
    return client.post(f"/api/schedules/{sid}/revisions", json={
        "code": code, "status": "S2 - Suitable for Information", **kw,
    }).json()


def issue(client, sid, revision_id):
    return client.post(
        f"/api/schedules/{sid}/revisions/{revision_id}/issue"
    ).json()


class TestIssuing:
    def test_a_new_revision_is_not_issued(self, client, schedule):
        revisions = add_revision(client, schedule)
        assert revisions[0]["issued"] is False
        assert revisions[0]["issued_at"] is None

    def test_issuing_freezes_it(self, client, schedule):
        rev = add_revision(client, schedule)[0]
        issued = issue(client, schedule, rev["id"])[0]
        assert issued["issued"] is True
        assert issued["issued_at"] is not None

    def test_the_snapshot_holds_the_computed_values(self, client, schedule):
        rev = add_revision(client, schedule)[0]
        issue(client, schedule, rev["id"])

        snap = client.get(
            f"/api/schedules/{schedule}/revisions/{rev['id']}/snapshot"
        ).json()
        row = snap["rows"][0]
        assert row["computed"]["Total Airflow (l/s)"] == 900
        assert row["computed"]["Manufacturer"] == "Systemair"
        assert snap["docnum"]

    def test_a_library_correction_cannot_change_an_issued_document(
        self, client, schedule
    ):
        """The whole point. Correct the library afterwards; the snapshot holds."""
        rev = add_revision(client, schedule)[0]
        issue(client, schedule, rev["id"])

        entries = client.get("/api/library/MVHR").json()
        client.put(f"/api/library/{entries[0]['id']}", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Vent-Axia", "Length (mm)": 9999},
        }).raise_for_status()

        live = client.get(f"/api/schedules/{schedule}").json()
        assert live["rows"][0]["computed"]["Manufacturer"] == "Vent-Axia"

        snap = client.get(
            f"/api/schedules/{schedule}/revisions/{rev['id']}/snapshot"
        ).json()
        assert snap["rows"][0]["computed"]["Manufacturer"] == "Systemair"
        assert snap["rows"][0]["computed"]["Length (mm)"] == 1200

    def test_issuing_twice_does_not_overwrite_the_record(self, client, schedule):
        rev = add_revision(client, schedule)[0]
        issue(client, schedule, rev["id"])
        first = client.get(
            f"/api/schedules/{schedule}/revisions/{rev['id']}/snapshot"
        ).json()["taken_at"]

        client.put(f"/api/schedules/{schedule}/rows/"
                   f"{client.get(f'/api/schedules/{schedule}').json()['rows'][0]['id']}",
                   json={"values": {"Unit Reference": "CHANGED"}})
        issue(client, schedule, rev["id"])

        again = client.get(
            f"/api/schedules/{schedule}/revisions/{rev['id']}/snapshot"
        ).json()
        assert again["taken_at"] == first
        assert again["rows"][0]["values"]["Unit Reference"] == "MVHR-01"

    def test_an_unissued_revision_has_no_snapshot_to_read(self, client, schedule):
        rev = add_revision(client, schedule)[0]
        response = client.get(
            f"/api/schedules/{schedule}/revisions/{rev['id']}/snapshot"
        )
        assert response.status_code == 404
        assert "has not been issued" in response.json()["detail"]


class TestIssuedRevisionsAreProtected:
    def test_an_issued_revision_cannot_be_deleted(self, client, schedule):
        rev = add_revision(client, schedule)[0]
        issue(client, schedule, rev["id"])
        response = client.delete(f"/api/schedules/{schedule}/revisions/{rev['id']}")
        assert response.status_code == 409
        assert "supersede it" in response.json()["detail"]

    def test_an_unissued_revision_can_still_be_deleted(self, client, schedule):
        rev = add_revision(client, schedule)[0]
        assert client.delete(
            f"/api/schedules/{schedule}/revisions/{rev['id']}"
        ).status_code == 200

    def test_an_issued_revisions_code_is_fixed(self, client, schedule):
        rev = add_revision(client, schedule)[0]
        issue(client, schedule, rev["id"])
        response = client.put(
            f"/api/schedules/{schedule}/revisions/{rev['id']}",
            json={"code": "P09", "status": "S2 - Suitable for Information"},
        )
        assert response.status_code == 409

    def test_other_details_of_an_issued_revision_stay_correctable(
        self, client, schedule
    ):
        rev = add_revision(client, schedule)[0]
        issue(client, schedule, rev["id"])
        response = client.put(
            f"/api/schedules/{schedule}/revisions/{rev['id']}",
            json={
                "code": rev["code"], "status": "S4 - Suitable for Stage Approval",
                "description": "typo fixed",
            },
        )
        assert response.status_code == 200


class TestDiff:
    def issue_two(self, client, schedule):
        p01 = add_revision(client, schedule, "P01")[0]
        issue(client, schedule, p01["id"])
        return p01

    def test_a_changed_duty_is_reported(self, client, schedule):
        p01 = self.issue_two(client, schedule)

        row = client.get(f"/api/schedules/{schedule}").json()["rows"][0]["id"]
        client.put(f"/api/schedules/{schedule}/rows/{row}", json={"values": {
            "Unit Reference": "MVHR-01", "Model Reference": "SYS-1",
            "Supply Airflow (l/s)": 600, "Extract Airflow (l/s)": 450,
            "Total Power Input (W)": 396,
        }})
        p02 = [r for r in add_revision(client, schedule, "P02") if r["code"] == "P02"][0]
        issue(client, schedule, p02["id"])

        diff = client.get(
            f"/api/schedules/{schedule}/revisions/{p01['id']}/diff?against={p02['id']}"
        ).json()

        assert diff["from"] == "P01" and diff["to"] == "P02"
        changed = {f["column"]: f for row in diff["changed"] for f in row["fields"]}
        assert changed["Supply Airflow (l/s)"]["before"] == 450
        assert changed["Supply Airflow (l/s)"]["after"] == 600
        assert changed["Total Airflow (l/s)"]["after"] == 1050, "derived changes are shown"

    def test_a_library_correction_shows_as_a_change(self, client, schedule):
        """Nobody retyped anything, but the document says something different."""
        p01 = self.issue_two(client, schedule)

        entries = client.get("/api/library/MVHR").json()
        client.put(f"/api/library/{entries[0]['id']}", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Vent-Axia", "Length (mm)": 1200},
        })

        diff = client.get(f"/api/schedules/{schedule}/revisions/{p01['id']}/diff").json()
        changed = {f["column"] for row in diff["changed"] for f in row["fields"]}
        assert "Manufacturer" in changed

    def test_added_and_removed_rows_are_reported(self, client, schedule):
        p01 = self.issue_two(client, schedule)
        client.post(f"/api/schedules/{schedule}/rows", json={
            "values": {"Unit Reference": "MVHR-02"},
        })
        diff = client.get(f"/api/schedules/{schedule}/revisions/{p01['id']}/diff").json()
        assert "MVHR-02" in diff["added"]

    def test_no_change_reports_nothing(self, client, schedule):
        p01 = self.issue_two(client, schedule)
        diff = client.get(f"/api/schedules/{schedule}/revisions/{p01['id']}/diff").json()
        assert diff["changed"] == [] and diff["added"] == [] and diff["removed"] == []
        assert diff["unchanged"] == 1

    def test_comparing_an_unissued_revision_is_refused(self, client, schedule):
        rev = add_revision(client, schedule)[0]
        response = client.get(f"/api/schedules/{schedule}/revisions/{rev['id']}/diff")
        assert response.status_code == 400
        assert "never issued" in response.json()["detail"]


class TestBulkRevision:
    @pytest.fixture()
    def project(self, client):
        project = client.post("/api/projects", json={
            "name": "J", "number": "CM2", "client": "C",
        }).json()
        building = project["buildings"][0]["id"]
        for code in ("MVHR", "AHU", "FCU"):
            client.post(
                f"/api/projects/{project['id']}/buildings/{building}/schedules",
                json={"code": code},
            )
        return client.get(f"/api/projects/{project['id']}").json()

    def test_preview_lists_what_would_change(self, client, project):
        result = client.post(f"/api/projects/{project['id']}/revisions/bulk", json={
            "schedule_ids": [], "status": "S2 - Suitable for Information",
        }).json()
        assert result["applied"] == 0
        assert len(result["changes"]) == 3
        assert all(c["to"] == "P01" for c in result["changes"])

    def test_each_schedule_continues_its_own_series(self, client, project):
        schedules = project["buildings"][0]["schedules"]
        # Put one schedule ahead of the others.
        for code in ("P01", "P02"):
            client.post(f"/api/schedules/{schedules[0]['id']}/revisions", json={"code": code})

        result = client.post(f"/api/projects/{project['id']}/revisions/bulk", json={
            "schedule_ids": [], "apply": True,
        }).json()
        by_code = {c["code"]: c for c in result["changes"]}
        assert by_code[schedules[0]["code"]]["to"] == "P03", "already at P02"
        assert by_code[schedules[1]["code"]]["to"] == "P01", "starting out"

    def test_applying_adds_the_revision_to_each(self, client, project):
        client.post(f"/api/projects/{project['id']}/revisions/bulk", json={
            "schedule_ids": [], "apply": True,
            "status": "S2 - Suitable for Information", "description": "Stage 4 issue",
        })
        for schedule in project["buildings"][0]["schedules"]:
            revisions = client.get(f"/api/schedules/{schedule['id']}/revisions").json()
            assert [r["code"] for r in revisions] == ["P01"]
            assert revisions[0]["description"] == "Stage 4 issue"

    def test_it_can_issue_at_the_same_time(self, client, project):
        client.post(f"/api/projects/{project['id']}/revisions/bulk", json={
            "schedule_ids": [], "apply": True, "issue": True,
        })
        for schedule in project["buildings"][0]["schedules"]:
            revisions = client.get(f"/api/schedules/{schedule['id']}/revisions").json()
            assert revisions[0]["issued"] is True

    def test_a_subset_can_be_selected(self, client, project):
        schedules = project["buildings"][0]["schedules"]
        result = client.post(f"/api/projects/{project['id']}/revisions/bulk", json={
            "schedule_ids": [schedules[0]["id"]], "apply": True,
        }).json()
        assert result["applied"] == 1
        assert client.get(f"/api/schedules/{schedules[1]['id']}/revisions").json() == []


class TestIssuedExport:
    def test_a_past_revision_exports_from_its_snapshot(self, client, schedule, tmp_path):
        from openpyxl import load_workbook

        rev = add_revision(client, schedule)[0]
        issue(client, schedule, rev["id"])

        entries = client.get("/api/library/MVHR").json()
        client.put(f"/api/library/{entries[0]['id']}", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "CHANGED-LATER"},
        })

        response = client.get(
            f"/api/schedules/{schedule}/revisions/{rev['id']}/export.xlsx"
        )
        assert response.status_code == 200
        path = tmp_path / "issued.xlsx"
        path.write_bytes(response.content)

        sheet = load_workbook(path)["Schedule"]
        header = [c.value for c in sheet[4]]
        manufacturer = sheet.cell(6, header.index("Manufacturer") + 1).value
        assert manufacturer == "Systemair", "the issued document keeps what it said"
