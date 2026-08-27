"""The whole stack, from the HTTP API to a recalculated workbook.

This is the test that matters most, because it exercises the seam the rewrite
was built around: a user types into the web grid, and the workbook they export
has to agree with what the grid showed them. The grid computes in Python and the
workbook computes in Excel, from one AST. If those two ever disagree the tool is
lying to somebody, so the assertion is made against a real spreadsheet engine
rather than against our own arithmetic.
"""

from __future__ import annotations

import csv
import subprocess
import tempfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from schedul.export import pdf

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")

needs_soffice = pytest.mark.skipif(not pdf.available(), reason="LibreOffice is not installed")


@pytest.fixture()
def client(tmp_path, monkeypatch) -> TestClient:
    monkeypatch.setenv("SCHEDUL_DATA", str(tmp_path))
    monkeypatch.setenv("SCHEDUL_DATABASE_URL", f"sqlite:///{tmp_path / 'e2e.db'}")

    import schedul.db.session as session_module

    session_module.SessionLocal = None
    session_module.init_db(f"sqlite:///{tmp_path / 'e2e.db'}")

    from schedul.api.main import app

    return TestClient(app)


@pytest.fixture()
def project(client) -> dict:
    response = client.post("/api/projects", json={
        "name": "Head Office Refurbishment",
        "number": "CM4220",
        "client": "Northern Estates",
        "prepared_by": "AG", "checked_by": "LJ", "approved_by": "RS",
    })
    assert response.status_code == 201, response.text
    return response.json()


def recalculate(xlsx: Path) -> dict[str, list[list[str]]]:
    with tempfile.TemporaryDirectory() as out_dir, tempfile.TemporaryDirectory() as profile:
        subprocess.run(
            [
                pdf.soffice_path(), "--headless", "--norestore", "--nolockcheck",
                f"-env:UserInstallation=file://{profile}",
                "--convert-to",
                "csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,false,false,-1",
                "--outdir", out_dir, str(xlsx),
            ],
            capture_output=True, text=True, timeout=240, check=False,
        )
        sheets = {}
        for path in Path(out_dir).glob("*.csv"):
            name = path.stem.split("-", 1)[1] if "-" in path.stem else path.stem
            with path.open(newline="", encoding="utf-8") as fh:
                sheets[name] = list(csv.reader(fh))
        return sheets


class TestSetupFlow:
    def test_a_new_project_gets_one_hidden_building(self, project):
        assert len(project["buildings"]) == 1
        assert project["schedule_count"] == 0

    def test_the_catalogue_is_ready_immediately(self, client):
        types = client.get("/api/catalogue").json()
        assert {"MVHR", "AHU", "FCU", "RADPANEL"} <= {t["code"] for t in types}

    def test_adding_schedules_numbers_them_from_ten(self, client, project):
        building = project["buildings"][0]["id"]
        for code in ("MVHR", "FCU", "AHU"):
            response = client.post(
                f"/api/projects/{project['id']}/buildings/{building}/schedules",
                json={"code": code},
            )
            assert response.status_code == 201, response.text
        latest = response.json()
        assert [s["number"] for s in latest["buildings"][0]["schedules"]] == [10, 11, 12]

    def test_volume_follows_the_type_without_being_set(self, client, project):
        building = project["buildings"][0]["id"]
        for code in ("MVHR", "FCU"):
            client.post(
                f"/api/projects/{project['id']}/buildings/{building}/schedules",
                json={"code": code},
            )
        schedules = {
            s["code"]: s["docnum"]
            for s in client.get(f"/api/projects/{project['id']}").json()["buildings"][0]["schedules"]
        }
        assert "-5_7-" in schedules["MVHR"], "MVHR is ventilation"
        assert "-5_6-" in schedules["FCU"], "an FCU is heating and cooling"


class TestEditing:
    @pytest.fixture()
    def schedule(self, client, project) -> str:
        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        return result["buildings"][0]["schedules"][0]["id"]

    def test_derived_columns_compute_as_the_user_types(self, client, schedule):
        grid = client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "MVHR-01",
            "Supply Airflow (l/s)": "450",
            "Extract Airflow (l/s)": "450",
            "Total Power Input (W)": "396",
        }}).json()
        computed = grid["rows"][0]["computed"]
        assert computed["Total Airflow (l/s)"] == 900
        assert computed["Specific Fan Power (W/(l/s))"] == pytest.approx(0.88)

    def test_numbers_typed_as_text_are_stored_as_numbers(self, client, schedule):
        """Stored as text they would reach the workbook as text: left-aligned
        and ignored by SUM."""
        grid = client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "MVHR-01", "Supply Airflow (l/s)": "450",
        }}).json()
        assert grid["rows"][0]["values"]["Supply Airflow (l/s)"] == 450

    def test_a_reference_with_a_leading_zero_stays_text(self, client, schedule):
        grid = client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "0123",
        }}).json()
        assert grid["rows"][0]["values"]["Unit Reference"] == "0123"

    def test_computed_columns_submitted_by_a_client_are_discarded(self, client, schedule):
        grid = client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "MVHR-01",
            "Supply Airflow (l/s)": 450,
            "Total Airflow (l/s)": 999999,
            "Manufacturer": "FORGED",
        }}).json()
        stored = grid["rows"][0]["values"]
        assert "Total Airflow (l/s)" not in stored
        assert "Manufacturer" not in stored

    def test_equipment_saved_once_populates_the_library_columns(self, client, schedule):
        client.post("/api/library", json={
            "type_code": "MVHR",
            "model_reference": "SYS-VSR-500",
            "values": {"Manufacturer": "Systemair", "Length (mm)": 1200},
        }).raise_for_status()

        grid = client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "MVHR-01", "Model Reference": "SYS-VSR-500",
        }}).json()
        computed = grid["rows"][0]["computed"]
        assert computed["Manufacturer"] == "Systemair"
        assert computed["Length (mm)"] == 1200

    def test_correcting_the_library_corrects_every_schedule_at_once(self, client, schedule):
        entry = client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Systemari"},
        }).json()
        client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "M-1", "Model Reference": "SYS-1",
        }})

        client.put(f"/api/library/{entry['id']}", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Systemair"},
        }).raise_for_status()

        grid = client.get(f"/api/schedules/{schedule}").json()
        assert grid["rows"][0]["computed"]["Manufacturer"] == "Systemair"

    def test_an_unknown_model_reference_is_explained_not_silently_blank(
        self, client, schedule
    ):
        grid = client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "M-1", "Model Reference": "NOT-A-PRODUCT",
        }}).json()
        problems = grid["rows"][0]["problems"]
        assert any("not in the equipment library" in p for p in problems.values())


class TestRegister:
    def test_the_register_reports_the_current_revision(self, client, project):
        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        schedule = result["buildings"][0]["schedules"][0]["id"]

        for code, status in [
            ("P01", "S2 - Suitable for Information"),
            ("P02", "S2 - Suitable for Information"),
            ("C01", "S4 - Suitable for Stage Approval"),
        ]:
            client.post(f"/api/schedules/{schedule}/revisions", json={
                "code": code, "status": status, "issue_date": "2026-06-09",
            }).raise_for_status()

        row = client.get("/api/register").json()[0]
        assert row["revision"] == "C01", "a published revision outranks every preliminary one"
        assert row["status"] == "S4"
        assert row["status_description"] == "Suitable for Stage Approval"
        assert row["file_name"].endswith(".xlsx")

    def test_the_next_revision_continues_its_own_series(self, client, project):
        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        schedule = result["buildings"][0]["schedules"][0]["id"]
        for code in ("P01", "P02", "P03"):
            client.post(f"/api/schedules/{schedule}/revisions", json={"code": code})

        assert client.get(f"/api/schedules/{schedule}/revisions/next").json()["code"] == "P04"
        assert (
            client.get(f"/api/schedules/{schedule}/revisions/next?published=true").json()["code"]
            == "C01"
        ), "the first published revision is C01, not C04"


class TestExport:
    @pytest.fixture()
    def populated(self, client, project) -> tuple[str, str]:
        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        schedule = result["buildings"][0]["schedules"][0]["id"]

        client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-VSR-500",
            "values": {"Manufacturer": "Systemair", "Length (mm)": 1200},
        })
        client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "MVHR-01",
            "Location": "Roof Plant Area",
            "Supply Airflow (l/s)": "450",
            "Extract Airflow (l/s)": "450",
            "Total Power Input (W)": "396",
            "Model Reference": "SYS-VSR-500",
        }})
        client.post(f"/api/schedules/{schedule}/revisions", json={
            "code": "P01", "status": "S2 - Suitable for Information",
            "issue_date": "2026-06-09",
        })
        return project["id"], schedule

    def test_the_export_carries_the_typed_data_as_numbers(self, client, populated, tmp_path):
        _, schedule = populated
        response = client.get(f"/api/schedules/{schedule}/export.xlsx")
        assert response.status_code == 200

        path = tmp_path / "export.xlsx"
        path.write_bytes(response.content)
        sheet = load_workbook(path)["Schedule"]

        assert sheet["A6"].value == "MVHR-01"
        assert sheet["D6"].value == 450 and isinstance(sheet["D6"].value, int)
        assert sheet["H6"].value == 396

    def test_a_project_exports_with_the_house_folder_layout(self, client, populated):
        project_id, _ = populated
        response = client.get(f"/api/projects/{project_id}/export.zip?fmt=xlsx")
        assert response.status_code == 200

        import io, zipfile

        names = zipfile.ZipFile(io.BytesIO(response.content)).namelist()
        # One building, so the files sit directly in Schedules/ (SPEC.md 4.3.1).
        assert any(n.startswith("Schedules/") and n.endswith(".xlsx") for n in names)
        assert "Schedules/MAINPROJECTINFO.xlsx" in names
        assert not any(n.count("/") > 1 for n in names)

    def test_a_multi_building_project_exports_into_per_building_folders(
        self, client, project
    ):
        first = project["buildings"][0]["id"]
        client.post(f"/api/projects/{project['id']}/buildings/{first}/schedules",
                    json={"code": "MVHR"})
        updated = client.post(f"/api/projects/{project['id']}/buildings",
                              json={"ref": "HQ014", "name": "East Wing"}).json()
        second = [b for b in updated["buildings"] if b["ref"] == "HQ014"][0]["id"]
        client.post(f"/api/projects/{project['id']}/buildings/{second}/schedules",
                    json={"code": "AHU"})

        import io, zipfile

        response = client.get(f"/api/projects/{project['id']}/export.zip?fmt=xlsx")
        names = zipfile.ZipFile(io.BytesIO(response.content)).namelist()
        folders = {n.split("/")[1] for n in names if n.count("/") > 1}
        assert "HQ014" in folders
        assert len(folders) == 2, "each building gets its own folder once there are several"

    @needs_soffice
    def test_the_workbook_agrees_with_the_grid(self, client, populated, tmp_path):
        """The seam: what the browser showed and what Excel computes must match."""
        _, schedule = populated

        grid = client.get(f"/api/schedules/{schedule}").json()
        computed = grid["rows"][0]["computed"]

        path = tmp_path / "export.xlsx"
        path.write_bytes(client.get(f"/api/schedules/{schedule}/export.xlsx").content)
        sheets = recalculate(path)
        excel = dict(zip(sheets["Schedule"][3], sheets["Schedule"][5]))

        assert float(excel["Total Airflow"]) == float(computed["Total Airflow (l/s)"])
        assert float(excel["Specific Fan Power"]) == pytest.approx(
            float(computed["Specific Fan Power (W/(l/s))"])
        )
        # And the library lookup resolved inside Excel, not just in our grid.
        assert excel["Manufacturer"] == computed["Manufacturer"] == "Systemair"

    @needs_soffice
    def test_the_pdf_is_produced(self, client, populated, tmp_path):
        _, schedule = populated
        response = client.get(f"/api/schedules/{schedule}/export.pdf")
        assert response.status_code == 200
        assert response.content[:5] == b"%PDF-"
        assert len(response.content) > 5000


class TestGridOperations:
    """Phase 1 editing: duplicate, paste modes and fill-down."""

    @pytest.fixture()
    def schedule(self, client, project) -> str:
        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        return result["buildings"][0]["schedules"][0]["id"]

    def add(self, client, schedule, **values):
        return client.post(
            f"/api/schedules/{schedule}/rows", json={"values": values}
        ).json()

    def test_duplicate_copies_the_row_and_inserts_it_below(self, client, schedule):
        self.add(client, schedule, **{"Unit Reference": "MVHR-01", "Supply Airflow (l/s)": 450})
        self.add(client, schedule, **{"Unit Reference": "MVHR-99"})
        grid = client.get(f"/api/schedules/{schedule}").json()
        first = grid["rows"][0]["id"]

        grid = client.post(f"/api/schedules/{schedule}/rows/{first}/duplicate").json()
        refs = [r["values"].get("Unit Reference") for r in grid["rows"]]
        assert refs == ["MVHR-01", "MVHR-01", "MVHR-99"], "the copy sits directly below"
        assert grid["rows"][1]["values"]["Supply Airflow (l/s)"] == 450
        assert [r["position"] for r in grid["rows"]] == [0, 1, 2]

    def test_paste_append_keeps_existing_rows(self, client, schedule):
        self.add(client, schedule, **{"Unit Reference": "MVHR-01"})
        grid = client.post(f"/api/schedules/{schedule}/rows/paste", json={
            "mode": "append",
            "rows": [{"values": {"Unit Reference": "MVHR-02"}}],
        }).json()
        assert [r["values"]["Unit Reference"] for r in grid["rows"]] == ["MVHR-01", "MVHR-02"]

    def test_paste_insert_puts_rows_at_a_position(self, client, schedule):
        for ref in ("A", "C"):
            self.add(client, schedule, **{"Unit Reference": ref})
        grid = client.post(f"/api/schedules/{schedule}/rows/paste", json={
            "mode": "insert", "position": 1,
            "rows": [{"values": {"Unit Reference": "B"}}],
        }).json()
        assert [r["values"]["Unit Reference"] for r in grid["rows"]] == ["A", "B", "C"]

    def test_paste_replace_is_the_only_destructive_mode(self, client, schedule):
        self.add(client, schedule, **{"Unit Reference": "OLD"})
        grid = client.post(f"/api/schedules/{schedule}/rows/paste", json={
            "mode": "replace", "confirm": True,
            "rows": [{"values": {"Unit Reference": "NEW"}}],
        }).json()
        assert [r["values"]["Unit Reference"] for r in grid["rows"]] == ["NEW"]

    def test_replace_is_refused_until_it_is_confirmed(self, client, schedule):
        self.add(client, schedule, **{"Unit Reference": "OLD"})
        response = client.post(f"/api/schedules/{schedule}/rows/paste", json={
            "mode": "replace",
            "rows": [{"values": {"Unit Reference": "NEW"}}],
        })
        assert response.status_code == 409
        assert "1 row(s) that have been filled in" in response.json()["detail"]
        grid = client.get(f"/api/schedules/{schedule}").json()
        assert [r["values"]["Unit Reference"] for r in grid["rows"]] == ["OLD"], (
            "a refused replace must leave the schedule exactly as it was"
        )

    def test_replacing_an_empty_schedule_needs_no_confirmation(self, client, schedule):
        response = client.post(f"/api/schedules/{schedule}/rows/paste", json={
            "mode": "replace",
            "rows": [{"values": {"Unit Reference": "NEW"}}],
        })
        assert response.status_code == 200, "there was nothing to lose"

    def test_pasted_numbers_are_still_coerced(self, client, schedule):
        grid = client.post(f"/api/schedules/{schedule}/rows/paste", json={
            "mode": "append",
            "rows": [{"values": {"Unit Reference": "A", "Supply Airflow (l/s)": "450"}}],
        }).json()
        assert grid["rows"][0]["values"]["Supply Airflow (l/s)"] == 450

    def test_fill_down_increments_a_reference(self, client, schedule):
        self.add(client, schedule, **{"Unit Reference": "MVHR-001"})
        for _ in range(3):
            self.add(client, schedule)

        grid = client.post(f"/api/schedules/{schedule}/rows/fill", json={
            "column": "Unit Reference", "start_position": 0, "mode": "series",
        }).json()
        assert [r["values"]["Unit Reference"] for r in grid["rows"]] == [
            "MVHR-001", "MVHR-002", "MVHR-003", "MVHR-004",
        ]

    def test_fill_down_copies_plain_text(self, client, schedule):
        self.add(client, schedule, **{"Unit Reference": "A", "Location": "Roof Plantroom"})
        self.add(client, schedule, **{"Unit Reference": "B"})

        grid = client.post(f"/api/schedules/{schedule}/rows/fill", json={
            "column": "Location", "start_position": 0,
        }).json()
        assert [r["values"].get("Location") for r in grid["rows"]] == [
            "Roof Plantroom", "Roof Plantroom",
        ]

    def test_fill_down_refuses_a_computed_column(self, client, schedule):
        self.add(client, schedule, **{"Unit Reference": "A"})
        self.add(client, schedule)
        response = client.post(f"/api/schedules/{schedule}/rows/fill", json={
            "column": "Total Airflow (l/s)", "start_position": 0,
        })
        assert response.status_code == 400
        assert "calculated" in response.json()["detail"]

    def test_fill_down_recomputes_derived_columns(self, client, schedule):
        self.add(client, schedule, **{
            "Unit Reference": "A", "Supply Airflow (l/s)": 450, "Extract Airflow (l/s)": 450,
        })
        self.add(client, schedule, **{"Unit Reference": "B", "Extract Airflow (l/s)": 450})

        grid = client.post(f"/api/schedules/{schedule}/rows/fill", json={
            "column": "Supply Airflow (l/s)", "start_position": 0, "mode": "copy",
        }).json()
        assert grid["rows"][1]["computed"]["Total Airflow (l/s)"] == 900


class TestColumnModelEndToEnd:
    """Visibility and overrides all the way to a recalculated workbook."""

    @pytest.fixture()
    def schedule(self, client, project) -> str:
        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        return result["buildings"][0]["schedules"][0]["id"]

    def test_a_project_column_appears_in_the_editor(self, client, project, schedule):
        client.put(f"/api/projects/{project['id']}/columns", json={
            "type_code": "MVHR",
            "columns": [{"kind": "input", "name": "Quantity", "width": 10, "example": 2}],
        }).raise_for_status()

        grid = client.get(f"/api/schedules/{schedule}").json()
        quantity = [c for c in grid["columns"] if c["name"] == "Quantity"]
        assert quantity and quantity[0]["project_extra"] is True
        assert quantity[0]["editable"] is True

    def test_a_project_column_holds_data_and_exports(self, client, project, schedule, tmp_path):
        client.put(f"/api/projects/{project['id']}/columns", json={
            "type_code": "MVHR",
            "columns": [{"kind": "input", "name": "Quantity", "width": 10, "example": 2}],
        })
        client.post(f"/api/schedules/{schedule}/rows", json={"values": {
            "Unit Reference": "MVHR-01", "Quantity": 3,
        }})

        grid = client.get(f"/api/schedules/{schedule}").json()
        assert grid["rows"][0]["values"]["Quantity"] == 3

        path = tmp_path / "x.xlsx"
        path.write_bytes(client.get(f"/api/schedules/{schedule}/export.xlsx").content)
        headers = [c.value for c in load_workbook(path)["Schedule"][4]]
        assert "Quantity" in headers

    def test_a_column_shadowing_a_base_column_is_refused(self, client, project):
        response = client.put(f"/api/projects/{project['id']}/columns", json={
            "type_code": "MVHR",
            "columns": [{"kind": "input", "name": "Location"}],
        })
        assert response.status_code == 400
        assert "already exists" in str(response.json()["detail"])

    def test_a_hidden_column_stays_out_of_the_workbook(self, client, project, schedule, tmp_path):
        client.put(f"/api/projects/{project['id']}/columns", json={
            "type_code": "MVHR",
            "columns": [{
                "kind": "input", "name": "Price", "width": 10,
                "visibility": {"xlsx": False, "pdf": False},
            }],
        }).raise_for_status()

        grid = client.get(f"/api/schedules/{schedule}").json()
        assert any(c["name"] == "Price" for c in grid["columns"]), "visible in the editor"

        path = tmp_path / "x.xlsx"
        path.write_bytes(client.get(f"/api/schedules/{schedule}/export.xlsx").content)
        headers = [c.value for c in load_workbook(path)["Schedule"][4]]
        assert "Price" not in headers, "internal data must not reach the deliverable"

    def test_an_override_survives_to_the_workbook_as_a_value(
        self, client, project, schedule, tmp_path
    ):
        client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Systemair", "Length (mm)": 1200},
        })
        grid = client.post(f"/api/schedules/{schedule}/rows", json={
            "values": {"Unit Reference": "M-1", "Model Reference": "SYS-1"},
            "overrides": {"Length (mm)": 1400},
        }).json()

        assert grid["rows"][0]["computed"]["Length (mm)"] == 1400
        assert grid["rows"][0]["computed"]["Manufacturer"] == "Systemair"
        assert grid["rows"][0]["overrides"] == {"Length (mm)": 1400}

        path = tmp_path / "x.xlsx"
        path.write_bytes(client.get(f"/api/schedules/{schedule}/export.xlsx").content)
        sheet = load_workbook(path)["Schedule"]
        header = [c.value for c in sheet[4]]
        length_col = header.index("Length") + 1
        manufacturer_col = header.index("Manufacturer") + 1

        # The overridden cell is a literal; leaving the formula would discard it.
        assert sheet.cell(6, length_col).value == 1400
        assert str(sheet.cell(6, manufacturer_col).value).startswith("=IF(")

    def test_clearing_an_override_restores_the_library_value(self, client, schedule):
        client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Length (mm)": 1200},
        })
        grid = client.post(f"/api/schedules/{schedule}/rows", json={
            "values": {"Unit Reference": "M-1", "Model Reference": "SYS-1"},
            "overrides": {"Length (mm)": 1400},
        }).json()
        row = grid["rows"][0]["id"]

        grid = client.put(f"/api/schedules/{schedule}/rows/{row}", json={
            "values": {"Unit Reference": "M-1", "Model Reference": "SYS-1"},
            "overrides": {"Length (mm)": ""},
        }).json()
        assert grid["rows"][0]["overrides"] == {}
        assert grid["rows"][0]["computed"]["Length (mm)"] == 1200

    @needs_soffice
    def test_the_workbook_still_agrees_with_the_grid_under_an_override(
        self, client, schedule, tmp_path
    ):
        """The seam again: an override must not make the two disagree."""
        client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Systemair"},
        })
        grid = client.post(f"/api/schedules/{schedule}/rows", json={
            "values": {
                "Unit Reference": "M-1", "Model Reference": "SYS-1",
                "Supply Airflow (l/s)": 500, "Extract Airflow (l/s)": 500,
                "Total Power Input (W)": 400,
            },
            "overrides": {"Manufacturer": "Vent-Axia"},
        }).json()

        path = tmp_path / "x.xlsx"
        path.write_bytes(client.get(f"/api/schedules/{schedule}/export.xlsx").content)
        sheets = recalculate(path)
        excel = dict(zip(sheets["Schedule"][3], sheets["Schedule"][5]))

        computed = grid["rows"][0]["computed"]
        assert excel["Manufacturer"] == computed["Manufacturer"] == "Vent-Axia"
        assert float(excel["Total Airflow"]) == float(computed["Total Airflow (l/s)"]) == 1000


class TestRegisterAndRooms:
    @pytest.fixture()
    def populated(self, client, project) -> str:
        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "RAD"},
        ).json()
        sid = result["buildings"][0]["schedules"][0]["id"]
        for ref, room in [("RAD-001", "RM8.64"), ("RAD-002", "RM8.64"), ("RAD-003", "RM2")]:
            client.post(f"/api/schedules/{sid}/rows", json={"values": {
                "Radiator Reference": ref, "Room Served": room,
            }})
        return project["id"]

    def test_rooms_group_equipment(self, client, populated):
        data = client.get(f"/api/projects/{populated}/rooms").json()
        rooms = {r["room"]: r for r in data["rooms"]}
        assert rooms["RM8.64"]["count"] == 2
        assert rooms["RM8.64"]["by_type"] == {"RAD": 2}
        assert rooms["RM2"]["count"] == 1

    def test_rooms_sort_naturally(self, client, populated):
        """RM2 before RM8.64, which plain string ordering gets backwards."""
        data = client.get(f"/api/projects/{populated}/rooms").json()
        assert [r["room"] for r in data["rooms"]] == ["RM2", "RM8.64"]

    def test_which_column_was_used_is_reported(self, client, populated):
        """A wrong guess about which column names a room should be visible."""
        data = client.get(f"/api/projects/{populated}/rooms").json()
        assert data["room_columns"]["RAD"] == "Room Served"

    def test_rows_with_no_room_are_counted_not_hidden(self, client, populated):
        schedule = client.get("/api/register").json()[0]["schedule_id"]
        client.post(f"/api/schedules/{schedule}/rows", json={
            "values": {"Radiator Reference": "RAD-099"},
        })
        data = client.get(f"/api/projects/{populated}/rooms").json()
        assert data["unassigned"] == 1

    def test_the_register_carries_what_search_needs(self, client, populated):
        row = client.get("/api/register").json()[0]
        for key in ("project_name", "project_number", "building", "code",
                    "document_number", "file_name", "status", "revision"):
            assert key in row


class TestNumberingRules:
    """P2.15: discipline from volume, and optional per-volume sequences."""

    def set_house(self, client, **changes):
        return client.put("/api/settings", json=changes)

    def test_discipline_follows_the_volume(self, client, project):
        """MVHR is ventilation so mechanical; EWH is domestic services so
        public health. Neither is set on the project."""
        building = project["buildings"][0]["id"]
        for code in ("MVHR", "EWH"):
            client.post(
                f"/api/projects/{project['id']}/buildings/{building}/schedules",
                json={"code": code},
            )
        by_code = {
            s["code"]: s["docnum"]
            for s in client.get(f"/api/projects/{project['id']}").json()["buildings"][0]["schedules"]
        }
        assert "-5_7-" in by_code["MVHR"] and "-M-" in by_code["MVHR"]
        assert "-5_3-" in by_code["EWH"] and "-P-" in by_code["EWH"]

    def test_a_project_override_still_wins_over_the_volume(self, client, project):
        """Resolution is schedule > building > type > project, so a project that
        needs one discipline throughout can still say so."""
        client.put(f"/api/projects/{project['id']}", json={
            "name": project["name"], "number": project["number"],
            "client": project["client"], "riba_stage": "Stage 4",
            "naming_overrides": {"discipline": "E"},
        })
        building = project["buildings"][0]["id"]
        client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "EWH"},
        )
        docnum = client.get(f"/api/projects/{project['id']}").json()[
            "buildings"][0]["schedules"][0]["docnum"]
        assert "-E-" in docnum and "-P-" not in docnum

    def test_clearing_the_lookup_leaves_discipline_project_scoped(self, client, project):
        self.set_house(client, volume_discipline={})
        building = project["buildings"][0]["id"]
        client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "EWH"},
        )
        docnum = client.get(f"/api/projects/{project['id']}").json()[
            "buildings"][0]["schedules"][0]["docnum"]
        assert "-M-" in docnum, "falls back to the project token"

    def test_numbering_is_per_building_by_default(self, client, project):
        building = project["buildings"][0]["id"]
        for code in ("MVHR", "EWH", "AHU"):
            client.post(
                f"/api/projects/{project['id']}/buildings/{building}/schedules",
                json={"code": code},
            )
        numbers = [
            s["number"]
            for s in client.get(f"/api/projects/{project['id']}").json()["buildings"][0]["schedules"]
        ]
        assert numbers == [10, 11, 12], "one sequence, whatever the volume"

    def test_per_volume_numbering_gives_each_volume_its_own_sequence(
        self, client, project
    ):
        self.set_house(client, numbering_scope="building_volume")
        building = project["buildings"][0]["id"]
        # MVHR and AHU are 5.7; EWH is 5.3.
        for code in ("MVHR", "EWH", "AHU"):
            client.post(
                f"/api/projects/{project['id']}/buildings/{building}/schedules",
                json={"code": code},
            )
        by_code = {
            s["code"]: s
            for s in client.get(f"/api/projects/{project['id']}").json()["buildings"][0]["schedules"]
        }
        assert by_code["MVHR"]["number"] == 10
        assert by_code["AHU"]["number"] == 11, "second in the 5.7 sequence"
        assert by_code["EWH"]["number"] == 10, "5.3 starts its own sequence"
        assert by_code["MVHR"]["docnum"] != by_code["EWH"]["docnum"]

    def test_a_retired_number_is_scoped_to_its_volume(self, client, project):
        self.set_house(client, numbering_scope="building_volume")
        building = project["buildings"][0]["id"]
        for code in ("MVHR", "AHU"):
            client.post(
                f"/api/projects/{project['id']}/buildings/{building}/schedules",
                json={"code": code},
            )
        schedules = client.get(f"/api/projects/{project['id']}").json()["buildings"][0]["schedules"]
        ahu = [s for s in schedules if s["code"] == "AHU"][0]
        client.delete(f"/api/projects/{project['id']}/schedules/{ahu['id']}")

        # 11 is retired in 5.7, so another ventilation type gets 12...
        client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "SUPGRILLE"},
        )
        # ...but 5.3 is untouched and still starts at 10.
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "EWH"},
        ).json()
        by_code = {s["code"]: s["number"] for s in result["buildings"][0]["schedules"]}
        assert by_code["SUPGRILLE"] == 12
        assert by_code["EWH"] == 10

    def test_the_start_number_is_configurable(self, client, project):
        """Numbering can begin at 1 without a code change."""
        settings = client.get("/api/settings").json()["house_standard"]
        naming = settings["naming"]
        naming["tokens"]["number"]["start"] = 1
        self.set_house(client, naming=naming)

        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        schedule = result["buildings"][0]["schedules"][0]
        assert schedule["number"] == 1
        assert "-00000001-" in schedule["docnum"]


class TestLibraryAudit:
    """P1.7: what changed in the library, and where it landed."""

    @pytest.fixture()
    def setup(self, client, project) -> dict:
        building = project["buildings"][0]["id"]
        result = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        sid = result["buildings"][0]["schedules"][0]["id"]
        entry = client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Systemari"}, "created_by": "AG",
        }).json()
        client.post(f"/api/schedules/{sid}/rows", json={"values": {
            "Unit Reference": "M-1", "Model Reference": "SYS-1",
        }})
        return {"schedule": sid, "entry": entry["id"]}

    def test_creating_an_entry_is_logged(self, client, setup):
        log = client.get("/api/library/review/changes").json()
        assert log[-1]["action"] == "created"
        assert log[-1]["model_reference"] == "SYS-1"
        assert log[-1]["actor"] == "AG"

    def test_an_edit_records_what_moved(self, client, setup):
        client.put(f"/api/library/{setup['entry']}", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Systemair"}, "created_by": "LJ",
        })
        change = client.get("/api/library/review/changes").json()[0]
        assert change["action"] == "updated"
        assert change["changes"] == [
            {"column": "Manufacturer", "before": "Systemari", "after": "Systemair"}
        ]

    def test_an_edit_that_changes_nothing_is_not_logged(self, client, setup):
        before = len(client.get("/api/library/review/changes").json())
        client.put(f"/api/library/{setup['entry']}", json={
            "type_code": "MVHR", "model_reference": "SYS-1",
            "values": {"Manufacturer": "Systemari"},
        })
        assert len(client.get("/api/library/review/changes").json()) == before

    def test_the_blast_radius_is_visible(self, client, setup):
        """Editing a product is not a local act; this is what it would touch."""
        impact = client.get(f"/api/library/{setup['entry']}/affected").json()
        assert impact["total_rows"] == 1
        assert impact["schedules"][0]["code"] == "MVHR"
        assert impact["schedules"][0]["rows"] == 1

    def test_an_overriding_row_is_counted_separately(self, client, setup):
        """A row that overrides the value would not move, so say so."""
        row = client.get(f"/api/schedules/{setup['schedule']}").json()["rows"][0]["id"]
        client.put(f"/api/schedules/{setup['schedule']}/rows/{row}", json={
            "values": {"Unit Reference": "M-1", "Model Reference": "SYS-1"},
            "overrides": {"Manufacturer": "Vent-Axia"},
        })
        impact = client.get(f"/api/library/{setup['entry']}/affected").json()
        assert impact["rows_overriding"] == 1

    def test_approving_is_logged(self, client, setup):
        client.post(f"/api/library/review/{setup['entry']}/approved")
        assert client.get("/api/library/review/changes").json()[0]["action"] == "approved"

    def test_the_log_can_be_filtered_by_type(self, client, setup):
        assert client.get("/api/library/review/changes?type_code=MVHR").json()
        assert client.get("/api/library/review/changes?type_code=AHU").json() == []
