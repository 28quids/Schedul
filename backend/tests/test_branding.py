"""Branding is a house standard, and an issued document is not a working file.

Two things are being defended here. One is that a practice's documents look like
they came from the same place, because the settings live on the organisation and
the renderer carries them out. The other is that a workbook a practice
configures can never come out with a broken reference in it -- which is why
hiding a field the cover reads is refused rather than obeyed.
"""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from schedul.core.branding import (
    COVER_FIELDS, REVISION_FIELDS, SAFE_FONTS, Branding, resolve_fields,
    validate_branding,
)

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")


class TestTheModel:
    def test_everything_shows_by_default(self):
        branding = Branding.from_dict({})
        assert [f.key for f in branding.cover_layout()] == [f.key for f in COVER_FIELDS]
        assert len(branding.revision_layout()) == len(REVISION_FIELDS)

    def test_an_optional_field_can_be_hidden(self):
        branding = Branding.from_dict({"cover_fields": {"building": False}})
        assert "building" not in [f.key for f in branding.cover_layout()]

    def test_a_field_the_workbook_reads_is_never_dropped(self):
        # Even asked for directly: resolve_fields is what the renderer calls,
        # and a hidden Revision row would leave the cover pointing at nothing.
        kept = resolve_fields(REVISION_FIELDS, {"revision": False, "date": False}, [])
        assert "revision" in [f.key for f in kept]
        assert "date" in [f.key for f in kept]

    def test_fields_can_be_reordered_and_the_rest_keep_their_places(self):
        branding = Branding.from_dict({"cover_order": ["revision", "date"]})
        order = [f.key for f in branding.cover_layout()]
        assert order[:2] == ["revision", "date"]
        assert set(order) == {f.key for f in COVER_FIELDS}

    def test_an_unknown_key_in_the_order_is_ignored_rather_than_breaking(self):
        branding = Branding.from_dict({"cover_order": ["nonsense", "date"]})
        assert [f.key for f in branding.cover_layout()][0] == "date"

    def test_colours_are_normalised_to_what_openpyxl_wants(self):
        branding = Branding.from_dict({"palette": {"accent": "#ff0000"}})
        assert branding.rgb("accent") == "FF0000"
        assert branding.colour("accent") == "FFFF0000"

    def test_an_eight_digit_house_style_colour_is_accepted(self):
        assert Branding.from_dict({"palette": {"title": "FF4D4D4D"}}).rgb("title") == "4D4D4D"

    def test_branding_carries_the_fonts_into_the_house_style(self):
        overrides = Branding.from_dict({"cover_font": "Calibri"}).house_style_overrides()
        assert overrides["cover_font"] == "Calibri"
        assert overrides["title_grey"].startswith("FF")


class TestValidation:
    def test_a_font_not_everybody_has_is_refused(self):
        problems = validate_branding({"cover_font": "Comic Sans MS"})
        assert any("every machine has" in p for p in problems)

    def test_every_offered_font_passes(self):
        for font in SAFE_FONTS:
            assert validate_branding({"cover_font": font, "schedule_font": font}) == []

    def test_a_bad_colour_is_refused(self):
        assert validate_branding({"palette": {"accent": "blue"}})

    def test_hiding_a_field_the_workbook_reads_is_refused_with_a_reason(self):
        problems = validate_branding({"revision_fields": {"document_number": False}})
        assert any("cannot be hidden" in p for p in problems)

    def test_an_unknown_field_is_refused(self):
        assert validate_branding({"cover_fields": {"colour_scheme": False}})

    def test_absurd_sizes_are_refused(self):
        assert validate_branding({"title_size": 400})
        assert validate_branding({"logo_scale": 40})


@pytest.fixture()
def client(tmp_path, monkeypatch) -> TestClient:
    monkeypatch.setenv("SCHEDUL_DATABASE_URL", f"sqlite:///{tmp_path / 'brand.db'}")
    import schedul.db.session as session_module

    session_module.SessionLocal = None
    session_module.init_db(f"sqlite:///{tmp_path / 'brand.db'}")
    from schedul.api.main import app

    return TestClient(app)


@pytest.fixture()
def schedule(client) -> str:
    project = client.post("/api/projects", json={
        "number": "CM1", "name": "Head Office", "client": "Northern Estates",
    }).json()
    building = project["buildings"][0]["id"]
    result = client.post(
        f"/api/projects/{project['id']}/buildings/{building}/schedules",
        json={"code": "MVHR"},
    ).json()
    sid = result["buildings"][0]["schedules"][0]["id"]
    client.post(f"/api/schedules/{sid}/rows", json={"values": {
        "Unit Reference": "MVHR-01", "Supply Airflow (l/s)": 450,
    }})
    return sid


def workbook(client, schedule, **params):
    query = "&".join(f"{k}={v}" for k, v in params.items())
    response = client.get(f"/api/schedules/{schedule}/export.xlsx" + (f"?{query}" if query else ""))
    assert response.status_code == 200, response.text
    return load_workbook(io.BytesIO(response.content))


def labels(sheet, column="A", top=1, bottom=60):
    return [
        sheet.cell(r, 1).value for r in range(top, bottom)
        if isinstance(sheet.cell(r, 1).value, str)
    ]


class TestThroughTheApi:
    def test_the_settings_screen_is_told_what_can_be_configured(self, client):
        data = client.get("/api/settings/branding").json()
        assert set(data["fonts"]) == set(SAFE_FONTS)
        fixed = [f for f in data["revision_fields"] if not f["optional"]]
        assert {"revision", "date", "document_number"} <= {f["key"] for f in fixed}

    def test_branding_is_saved_and_read_back(self, client):
        client.put("/api/settings", json={"branding": {
            "cover_font": "Calibri", "palette": {"accent": "112233"},
            "cover_subtitle": "Mechanical Services",
        }})
        branding = client.get("/api/settings/branding").json()["branding"]
        assert branding["cover_font"] == "Calibri"
        assert branding["palette"]["accent"] == "112233"

    def test_a_refused_change_leaves_the_branding_alone(self, client):
        response = client.put("/api/settings", json={"branding": {"cover_font": "Wingdings"}})
        assert response.status_code == 400
        assert client.get("/api/settings/branding").json()["branding"]["cover_font"] == "Verdana"

    def test_the_preview_says_what_a_change_would_produce_without_saving(self, client):
        preview = client.post("/api/settings/branding/preview", json={"branding": {
            "cover_fields": {"building": False},
        }}).json()
        assert "building" not in [f["key"] for f in preview["preview"]["cover"]]
        assert client.get("/api/settings/branding").json()["preview"]["cover"], "unchanged"
        assert "building" in [
            f["key"] for f in client.get("/api/settings/branding").json()["preview"]["cover"]
        ]

    def test_the_preview_reports_problems_rather_than_refusing(self, client):
        preview = client.post("/api/settings/branding/preview", json={
            "branding": {"cover_font": "Wingdings"},
        }).json()
        assert preview["problems"]


class TestItReachesTheDocument:
    def test_the_chosen_font_is_what_the_cover_is_set_in(self, client, schedule):
        client.put("/api/settings", json={"branding": {"cover_font": "Georgia"}})
        cover = workbook(client, schedule)["Front Cover"]
        title = next(
            cover.cell(r, 1) for r in range(1, 60)
            if cover.cell(r, 1).value and str(cover.cell(r, 1).value).startswith("=Config")
        )
        assert title.font.name == "Georgia"

    def test_the_accent_colour_reaches_the_title(self, client, schedule):
        client.put("/api/settings", json={"branding": {"palette": {"accent": "FF0000"}}})
        cover = workbook(client, schedule)["Front Cover"]
        titles = [
            cover.cell(r, 1) for r in range(1, 60)
            if cover.cell(r, 1).value and "MECHANICAL" in str(cover.cell(r, 1).value).upper()
        ]
        # openpyxl normalises a colour to eight digits; the last six are the ones
        # that carry the meaning.
        assert titles and titles[0].font.color.rgb[-6:] == "FF0000"

    def test_hiding_the_building_takes_it_off_both_pages(self, client, schedule):
        before = workbook(client, schedule)
        assert "Building" in labels(before["Revision page"])

        client.put("/api/settings", json={"branding": {
            "cover_fields": {"building": False},
            "revision_fields": {"building": False},
        }})
        after = workbook(client, schedule)
        assert "Building" not in labels(after["Revision page"])
        assert "Building" not in labels(after["Front Cover"])

    def test_a_hidden_field_does_not_leave_a_gap_in_the_summary(self, client, schedule):
        from schedul.core.branding import REVISION_FIELDS

        client.put("/api/settings", json={"branding": {
            "revision_fields": {"building": False, "bsuid": False},
        }})
        page = workbook(client, schedule)["Revision page"]

        # The block runs from row 10 to the first blank row. Two fields were
        # hidden, so it is two rows shorter -- not the same length with holes in.
        block = []
        row = 10
        while page.cell(row, 1).value:
            block.append(page.cell(row, 1).value)
            row += 1
        assert len(block) == len(REVISION_FIELDS) - 2
        assert "Building" not in block and "BSUID" not in block

    def test_reordering_the_cover_moves_the_fields(self, client, schedule):
        client.put("/api/settings", json={"branding": {
            "cover_order": ["revision", "recipient"],
        }})
        cover = workbook(client, schedule)["Front Cover"]
        found = [v for v in labels(cover) if v in ("Revision", "Intended for")]
        assert found == ["Revision", "Intended for"]

    def test_the_cover_still_reads_the_revision_page_after_a_reorder(self, client, schedule):
        client.put("/api/settings", json={"branding": {
            "revision_order": ["schedule_name", "document_number", "revision"],
        }})
        book = workbook(client, schedule)
        page, cover = book["Revision page"], book["Front Cover"]

        # Whatever row 'Document no' ended up on, the cover points at that row.
        row = next(r for r in range(1, 40) if page.cell(r, 1).value == "Document no")
        formulas = [
            str(cover.cell(r, 1).value) for r in range(1, 60)
            if cover.cell(r, 1).value and str(cover.cell(r, 1).value).startswith("=")
        ]
        assert any(f"'Revision page'!B{row}" in f for f in formulas)

    def test_the_subtitle_and_footer_slots_are_written(self, client, schedule):
        client.put("/api/settings", json={"branding": {
            "cover_subtitle": "Mechanical Services",
            "cover_footer": "Bovis Engineering Ltd · Registered in England",
        }})
        values = labels(workbook(client, schedule)["Front Cover"], bottom=70)
        assert "Mechanical Services" in values
        assert any("Registered in England" in v for v in values)


class TestIssueThemeVersusEditorTheme:
    def test_an_export_is_an_issued_document_by_default(self, client, schedule):
        sheet = workbook(client, schedule)["Schedule"]
        cell = sheet.cell(6, 1)
        assert cell.fill.fgColor.rgb[-6:] == "FFFFFF", (
            "an issued document must not carry the yellow editing fill"
        )
        assert cell.font.color is None or cell.font.color.rgb[-6:] != "0000FF", (
            "nor the blue input text colour"
        )

    def test_the_working_copy_keeps_the_editing_colours(self, client, schedule):
        sheet = workbook(client, schedule, theme="editor")["Schedule"]
        cell = sheet.cell(6, 1)
        assert cell.fill.fgColor.rgb[-6:] == "FFFFCC"
        assert cell.font.color.rgb[-6:] == "0000FF"

    def test_an_unknown_theme_is_refused_rather_than_guessed_at(self, client, schedule):
        response = client.get(f"/api/schedules/{schedule}/export.xlsx?theme=fancy")
        assert response.status_code == 400

    def test_the_two_themes_hold_the_same_numbers(self, client, schedule):
        issued = workbook(client, schedule)["Schedule"]
        working = workbook(client, schedule, theme="editor")["Schedule"]
        for row in range(6, 9):
            for col in range(1, 12):
                assert issued.cell(row, col).value == working.cell(row, col).value, (
                    "the theme changes how it looks, never what it says"
                )


class TestOneJobDifferingFromThePractice:
    """A project answering "which fields does a document carry" for itself.

    The house standard's own note used to say "hide what a job does not need",
    which was not true of a setting that reached every job in the practice. This
    is the setting that makes it true.
    """

    def test_a_project_follows_the_practice_until_it_says_otherwise(self, client):
        project = client.post("/api/projects", json={"number": "CM1"}).json()
        data = client.get(f"/api/projects/{project['id']}/branding").json()
        assert data["overrides"] == {}
        assert all(f["project"] is None for f in data["cover_fields"])
        assert "building" in [f["key"] for f in data["preview"]["cover"]]

    def test_hiding_a_field_on_one_job_leaves_every_other_job_alone(self, client, schedule):
        first = client.get(f"/api/schedules/{schedule}").json()["project_id"]
        second = client.post("/api/projects", json={"number": "CM2"}).json()

        client.put(f"/api/projects/{first}/branding", json={
            "cover_fields": {"building": False},
            "revision_fields": {"building": False},
        })

        mine = client.get(f"/api/projects/{first}/branding").json()
        theirs = client.get(f"/api/projects/{second['id']}/branding").json()
        assert "building" not in [f["key"] for f in mine["preview"]["cover"]]
        assert "building" in [f["key"] for f in theirs["preview"]["cover"]]
        assert client.get("/api/settings/branding").json()["branding"]["cover_fields"] == {}

    def test_it_reaches_the_workbook(self, client, schedule):
        project = client.get(f"/api/schedules/{schedule}").json()["project_id"]
        assert "Building" in labels(workbook(client, schedule)["Revision page"])

        client.put(f"/api/projects/{project}/branding", json={
            "cover_fields": {"building": False},
            "revision_fields": {"building": False},
        })
        after = workbook(client, schedule)
        assert "Building" not in labels(after["Revision page"])
        assert "Building" not in labels(after["Front Cover"])

    def test_a_project_can_show_what_the_practice_hides(self, client, schedule):
        project = client.get(f"/api/schedules/{schedule}").json()["project_id"]
        client.put("/api/settings", json={"branding": {"cover_fields": {"building": False}}})
        assert "Building" not in labels(workbook(client, schedule)["Front Cover"])

        client.put(f"/api/projects/{project}/branding", json={
            "cover_fields": {"building": True},
        })
        assert "Building" in labels(workbook(client, schedule)["Front Cover"])

    def test_a_project_cannot_hide_a_row_the_workbook_reads(self, client):
        project = client.post("/api/projects", json={"number": "CM1"}).json()
        response = client.put(f"/api/projects/{project['id']}/branding", json={
            "revision_fields": {"document_number": False},
        })
        assert response.status_code == 400
        assert "cannot be hidden" in response.text

    def test_a_project_cannot_change_the_practice_typeface(self, client, schedule):
        project = client.get(f"/api/schedules/{schedule}").json()["project_id"]
        client.put(f"/api/projects/{project}/branding", json={
            "cover_fields": {}, "cover_subtitle": "Mechanical Services",
        })
        stored = client.get(f"/api/projects/{project}/branding").json()["overrides"]
        assert "cover_font" not in stored, (
            "a per-project typeface would quietly end the house standard"
        )

    def test_following_the_practice_again_drops_the_overrides(self, client, schedule):
        project = client.get(f"/api/schedules/{schedule}").json()["project_id"]
        client.put(f"/api/projects/{project}/branding", json={
            "cover_fields": {"building": False},
        })
        client.put(f"/api/projects/{project}/branding", json={"cover_fields": {}})
        assert client.get(f"/api/projects/{project}/branding").json()["overrides"] == {}
        assert "Building" in labels(workbook(client, schedule)["Front Cover"])
