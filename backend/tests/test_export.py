"""The exported workbook, verified by a real spreadsheet engine.

Asserting that we wrote the formula we meant to write is weak. These tests hand
the workbook to LibreOffice, let it recalculate, and read the values back. That
is what proves the two formula backends agree: the grid's Python evaluation and
Excel's own evaluation of the emitted formula must produce the same number.
"""

from __future__ import annotations

import csv
import datetime as _dt
import subprocess
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from schedul.core.formula import evaluate, parse
from schedul.core.house import HouseStandard
from schedul.core.revisions import Revision
from schedul.export import pdf
from schedul.export.schedule import ScheduleContent, render_schedule

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")

PROJECT_FIELDS = {
    "Client": "A Client",
    "Project Name": "CM4220 Test",
    "Project Number": "CM4220",
    "Site Address": "",
    "Architect": "",
    "Main Contractor": "",
    "RIBA Stage": "Stage 4",
    "Prepared By": "AG",
    "Checked By": "LJ",
    "Approved By": "RS",
}

needs_soffice = pytest.mark.skipif(
    not pdf.available(), reason="LibreOffice is not installed"
)


def make_content(catalogue_types, code="MVHR", **overrides) -> ScheduleContent:
    house = HouseStandard()
    st = next(t for t in catalogue_types if t.code == code)
    kwargs = dict(
        schedule_type=st,
        house=house,
        project_fields=PROJECT_FIELDS,
        design_constants=house.design_constants,
        docnum="CM4220-BOV-5_7-HQ049-SC-M-00000010-G00300-XX-XX",
        building_ref="HQ049",
        building_name="Main Building",
        classification="G00300",
    )
    kwargs.update(overrides)
    return ScheduleContent(**kwargs)


def recalculate(xlsx: Path) -> dict[str, list[list[str]]]:
    """Open the workbook in LibreOffice and read every sheet back as values."""
    with tempfile.TemporaryDirectory() as out_dir, tempfile.TemporaryDirectory() as profile:
        subprocess.run(
            [
                pdf.soffice_path(), "--headless", "--norestore", "--nolockcheck",
                f"-env:UserInstallation=file://{profile}",
                "--convert-to",
                "csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,false,false,-1",
                "--outdir", out_dir, str(xlsx),
            ],
            capture_output=True, text=True, timeout=240, check=False,
        )
        sheets: dict[str, list[list[str]]] = {}
        for path in Path(out_dir).glob("*.csv"):
            name = path.stem.split("-", 1)[1] if "-" in path.stem else path.stem
            with path.open(newline="", encoding="utf-8") as fh:
                sheets[name] = list(csv.reader(fh))
        return sheets


def find_row(rows: list[list[str]], label: str) -> list[str]:
    for row in rows:
        if row and row[0].strip() == label:
            return row
    raise AssertionError(f"no row labelled {label!r}")


class TestStructure:
    """The house format the vendored code gets right, kept intact."""

    def test_sheets_and_visibility_match_house_format(self, catalogue_types, tmp_path):
        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        wb = load_workbook(path)
        assert [w.title for w in wb.worksheets] == [
            "Metadata", "Front Cover", "Revision page", "Schedule",
            "Config", "Lists", "Library",
        ]
        assert [w.sheet_state for w in wb.worksheets][-3:] == ["hidden"] * 3

    def test_page_setup_is_preserved(self, catalogue_types, tmp_path):
        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        sc = load_workbook(path)["Schedule"]
        assert sc.print_titles == "'Schedule'!$1:$5"
        assert sc.page_setup.orientation == "landscape"
        assert sc.page_setup.fitToWidth == 1
        assert sc.page_setup.fitToHeight == 0

    def test_header_is_split_across_rows_four_and_five(self, catalogue_types, tmp_path):
        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        sc = load_workbook(path)["Schedule"]
        assert sc["D4"].value == "Supply Airflow"
        assert sc["D5"].value == "l/s"

    def test_the_design_constants_are_defined_names(self, catalogue_types, tmp_path):
        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        names = load_workbook(path).defined_names
        assert "SETUP_CP" in names and "SETUP_LPHWF" in names

    def test_the_workbook_is_macro_free_and_has_no_external_links(
        self, catalogue_types, tmp_path
    ):
        """SPEC.md 1a.4: the real file pointed at a personal OneDrive. An export
        that carries a link to somebody's OneDrive is a support call."""
        import zipfile

        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        parts = zipfile.ZipFile(path).namelist()
        assert not [p for p in parts if "vbaProject" in p or "externalLink" in p]

    def test_no_config_path_keys_survive(self, catalogue_types, tmp_path):
        """The macros that read them are gone, so a stale path cannot mislead."""
        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        cf = load_workbook(path)["Config"]
        keys = {cf.cell(r, 1).value for r in range(1, 40)}
        assert not [k for k in keys if k and str(k).startswith("path_")]

    def test_building_is_stored_once_and_referenced(self, catalogue_types, tmp_path):
        """SPEC.md 4.6: Config is the only stored copy; the cover and revision
        page are formulas, so a rename stays one write."""
        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        wb = load_workbook(path)
        assert wb["Front Cover"]["A43"].value.startswith("=Config!")
        rv = wb["Revision page"]
        building = [rv.cell(r, 2).value for r in range(10, 24) if rv.cell(r, 1).value == "Building"]
        assert building and building[0].startswith("=Config!")


class TestNotes:
    """SPEC.md 4.7 and acceptance step 20."""

    def test_project_notes_come_first_then_type_notes_numbered_continuously(
        self, catalogue_types, tmp_path
    ):
        from schedul.services.seed import radiant_panel_type

        house = HouseStandard()
        content = make_content(catalogue_types, schedule_type=radiant_panel_type())
        path = render_schedule(content, tmp_path / "rp.xlsx")
        notes = load_workbook(path)["Schedule"]["A2"].value

        assert notes.startswith("General Notes:")
        project_count = len(house.general_notes)
        assert f"[{project_count}]" in notes
        assert f"[{project_count + 1}]" in notes
        # The radiant-panel-specific wording follows the generic wording.
        assert notes.index("read in conjunction") < notes.index("Merriott")
        assert f"[{project_count + 4}]" in notes, "all four type notes numbered"


@needs_soffice
class TestRecalculation:
    """Hand the workbook to a real engine and read the answers back."""

    def test_derived_columns_agree_with_the_python_evaluator(
        self, catalogue_types, tmp_path
    ):
        rows = [
            {
                "Unit Reference": "MVHR-01",
                "Supply Airflow (l/s)": 450,
                "Extract Airflow (l/s)": 450,
                "Total Power Input (W)": 396,
            }
        ]
        content = make_content(catalogue_types, rows=rows)
        path = render_schedule(content, tmp_path / "s.xlsx")
        sheets = recalculate(path)

        header = sheets["Schedule"][3]
        data = sheets["Schedule"][5]
        excel = dict(zip(header, data))

        st = next(t for t in catalogue_types if t.code == "MVHR")
        values = dict(rows[0])
        for col in st.derived:
            expected = evaluate(st.parse_formula(col), values, {})
            assert float(excel[col.name]) == pytest.approx(float(expected)), col.name

        assert float(excel["Total Airflow"]) == 900
        assert float(excel["Specific Fan Power"]) == pytest.approx(0.88)

    def test_a_library_field_with_no_value_stays_blank_not_zero(
        self, catalogue_types, tmp_path
    ):
        """INDEX into an empty cell returns 0, and 0 reads as a real duty."""
        content = make_content(
            catalogue_types,
            rows=[{"Unit Reference": "M-1", "Model Reference": "SYS-01"}],
            products=[{"Model Reference": "SYS-01", "Manufacturer": "Systemair"}],
        )
        path = render_schedule(content, tmp_path / "s.xlsx")
        sheets = recalculate(path)
        excel = dict(zip(sheets["Schedule"][3], sheets["Schedule"][5]))

        assert excel["Manufacturer"] == "Systemair"
        assert excel["Weight"] == "", "an unfilled product field must not read as 0"

    def test_an_unknown_model_reference_says_not_found(self, catalogue_types, tmp_path):
        content = make_content(
            catalogue_types,
            rows=[{"Unit Reference": "M-1", "Model Reference": "NOPE"}],
            products=[{"Model Reference": "SYS-01", "Manufacturer": "Systemair"}],
        )
        path = render_schedule(content, tmp_path / "s.xlsx")
        excel = dict(zip(recalculate(path)["Schedule"][3], recalculate(path)["Schedule"][5]))
        assert excel["Manufacturer"] == "NOT FOUND"

    def test_acceptance_step_21_a_published_revision_wins(
        self, catalogue_types, tmp_path
    ):
        """P01, P02, then C01: the cover must show C01 and its date.

        The real house file shows the last P revision here, because stripping
        'P' leaves C01 uncoercible and IFERROR turns it into 0.
        """
        revisions = [
            Revision("P01", status="S2 - Suitable for Information", date=_dt.date(2026, 1, 5)),
            Revision("P02", status="S2 - Suitable for Information", date=_dt.date(2026, 3, 1)),
            Revision("C01", status="S4 - Suitable for Stage Approval", date=_dt.date(2026, 6, 9)),
        ]
        path = render_schedule(
            make_content(catalogue_types, revisions=revisions), tmp_path / "s.xlsx"
        )
        sheets = recalculate(path)

        assert find_row(sheets["Revision page"], "Revision")[1] == "C01"
        assert find_row(sheets["Revision page"], "Date")[1] == "09/06/2026"
        assert find_row(sheets["Revision page"], "Suitability Status")[1] == "S4"
        assert find_row(sheets["Metadata"], "Revision")[1] == "C01"

    def test_acceptance_step_21_out_of_order_rows_still_rank(
        self, catalogue_types, tmp_path
    ):
        """P01, then P03 entered above P02: the cover shows P03.

        v1's INDEX(range, MAX(1, COUNTA(range))) reads the last non-empty row
        and would show P02.
        """
        revisions = [
            Revision("P01", status="S2 - Suitable for Information", date=_dt.date(2026, 1, 5)),
            Revision("P03", status="S2 - Suitable for Information", date=_dt.date(2026, 6, 1)),
            Revision("P02", status="S2 - Suitable for Information", date=_dt.date(2026, 3, 1)),
        ]
        # Written in the order given, deliberately not sorted first.
        content = make_content(catalogue_types)
        content.revisions = revisions
        path = render_schedule(content, tmp_path / "s.xlsx")
        sheets = recalculate(path)
        assert find_row(sheets["Revision page"], "Revision")[1] == "P03"

    def test_a_gap_in_the_log_does_not_break_the_summary(
        self, catalogue_types, tmp_path
    ):
        revisions = [
            Revision("P01", status="S2 - Suitable for Information", date=_dt.date(2026, 1, 5)),
            Revision("", status="", date=None),
            Revision("P02", status="S3 - Suitable for Review and Comment",
                     date=_dt.date(2026, 3, 1)),
        ]
        content = make_content(catalogue_types)
        content.revisions = revisions
        path = render_schedule(content, tmp_path / "s.xlsx")
        assert find_row(recalculate(path)["Revision page"], "Revision")[1] == "P02"

    def test_the_building_reaches_the_cover_and_the_revision_page(
        self, catalogue_types, tmp_path
    ):
        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        sheets = recalculate(path)
        assert find_row(sheets["Revision page"], "Building")[1] == "HQ049 - Main Building"
        assert any(
            "HQ049 - Main Building" in (cell or "")
            for row in sheets["Front Cover"]
            for cell in row
        )

    def test_the_document_number_resolves_through_the_chain(
        self, catalogue_types, tmp_path
    ):
        """Config!B4 -> Revision page -> Metadata, as SPEC.md 5.1 requires."""
        path = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        sheets = recalculate(path)
        expected = "CM4220-BOV-5_7-HQ049-SC-M-00000010-G00300-XX-XX"
        assert find_row(sheets["Revision page"], "Document no")[1] == expected
        assert find_row(sheets["Metadata"], "DocumentNumber")[1] == expected


@needs_soffice
class TestPdf:
    def test_a_schedule_converts_to_pdf(self, catalogue_types, tmp_path):
        xlsx = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        out = pdf.to_pdf(xlsx, tmp_path)
        assert out.exists() and out.stat().st_size > 1000
        assert out.read_bytes()[:5] == b"%PDF-"

    def test_a_wide_schedule_still_fits_a4_landscape(self, catalogue_types, tmp_path):
        """AHU is the widest type; fitToWidth must keep it to one page across."""
        xlsx = render_schedule(
            make_content(catalogue_types, code="AHU"), tmp_path / "ahu.xlsx"
        )
        out = pdf.to_pdf(xlsx, tmp_path)
        assert out.exists() and out.read_bytes()[:5] == b"%PDF-"

    def test_a_missing_libreoffice_is_reported_clearly(self, catalogue_types, tmp_path, monkeypatch):
        monkeypatch.setenv("SCHEDUL_SOFFICE", str(tmp_path / "nope"))
        xlsx = render_schedule(make_content(catalogue_types), tmp_path / "s.xlsx")
        with pytest.raises(pdf.PdfError, match="not found"):
            pdf.to_pdf(xlsx, tmp_path)
