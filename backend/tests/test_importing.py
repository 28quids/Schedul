"""Bringing a supplier's range into the library, forty rows at a time.

The rule being defended is that nothing is written until it has been shown. A
careless mapping can overwrite a hundred correct values in one click, so the dry
run is the default and applying is the exception -- and applying carries out the
plan that was shown rather than working it out again.
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

pytestmark = pytest.mark.filterwarnings("ignore::DeprecationWarning")

BLOCK = (
    "Model Reference\tManufacturer\tLength (mm)\n"
    "SYS-VSR-500\tSystemair\t1200\n"
    "SYS-VSR-700\tSystemair\t1400\n"
)


@pytest.fixture()
def client(tmp_path, monkeypatch) -> TestClient:
    monkeypatch.setenv("SCHEDUL_DATABASE_URL", f"sqlite:///{tmp_path / 'import.db'}")
    import schedul.db.session as session_module

    session_module.SessionLocal = None
    session_module.init_db(f"sqlite:///{tmp_path / 'import.db'}")
    from schedul.api.main import app

    return TestClient(app)


def plan(client, **body):
    response = client.post("/api/library/import", json={"type_code": "MVHR", **body})
    assert response.status_code == 200, response.text
    return response.json()


def library(client, code="MVHR"):
    return {e["model_reference"]: e for e in client.get(f"/api/library/{code}").json()}


class TestPlanning:
    def test_the_plan_says_what_each_row_would_do(self, client):
        result = plan(client, text=BLOCK)
        assert result["counts"]["create"] == 2
        assert [r["model_reference"] for r in result["rows"]] == ["SYS-VSR-500", "SYS-VSR-700"]
        assert result["header_detected"] is True

    def test_planning_writes_nothing(self, client):
        plan(client, text=BLOCK)
        assert library(client) == {}

    def test_a_header_matches_the_columns_by_name_in_any_order(self, client):
        result = plan(client, text=(
            "Manufacturer\tModel Reference\n"
            "Systemair\tSYS-1\n"
        ))
        assert result["rows"][0]["values"]["Manufacturer"] == "Systemair"
        assert result["rows"][0]["model_reference"] == "SYS-1"

    def test_without_a_header_the_caller_maps_the_columns(self, client):
        result = plan(
            client,
            text="SYS-1\tSystemair\n",
            mapping=["Model Reference", "Manufacturer"],
        )
        assert result["rows"][0]["values"] == {"Manufacturer": "Systemair"}

    def test_an_unmapped_column_is_ignored_rather_than_guessed_at(self, client):
        result = plan(
            client,
            text="SYS-1\tred\n",
            mapping=["Model Reference", None],
        )
        assert result["rows"][0]["values"] == {}

    def test_a_column_that_is_not_a_library_field_cannot_be_mapped(self, client):
        # 'Supply Airflow' is an input column: it differs per unit, so a value
        # imported into it would be a stale copy the moment it landed.
        offered = client.get("/api/library/MVHR/import/columns").json()["columns"]
        assert "Supply Airflow (l/s)" not in offered
        assert "Model Reference" in offered and "Manufacturer" in offered

    def test_without_a_model_reference_nothing_can_be_imported(self, client):
        result = plan(client, text="Systemair\t1200\n", mapping=["Manufacturer", "Length (mm)"])
        assert result["rows"] == []
        assert any("Model Reference" in w for w in result["warnings"])
        assert result["can_apply"] is False

    def test_a_row_with_a_blank_reference_is_refused_with_a_reason(self, client):
        result = plan(client, text="Model Reference\tManufacturer\n\tSystemair\n")
        assert result["rows"][0]["action"] == "skip"
        assert "nothing to key this on" in result["rows"][0]["reason"]

    def test_the_same_reference_twice_in_one_paste_is_caught(self, client):
        result = plan(client, text=(
            "Model Reference\tManufacturer\nSYS-1\tSystemair\nSYS-1\tSystemair AB\n"
        ))
        assert result["counts"]["create"] == 1
        assert "line 1" in result["rows"][1]["reason"]

    def test_an_empty_paste_says_so(self, client):
        result = plan(client, text="   ")
        assert any("nothing to import" in w for w in result["warnings"])


class TestDuplicatesAndUpdates:
    def test_an_existing_product_is_an_update_not_a_second_entry(self, client):
        plan(client, text=BLOCK, apply=True)
        result = plan(client, text=(
            "Model Reference\tManufacturer\tLength (mm)\nSYS-VSR-500\tSystemair\t1250\n"
        ))
        row = result["rows"][0]
        assert row["action"] == "update"
        assert row["changes"] == [{"column": "Length (mm)", "before": 1200, "after": "1250"}]

    def test_a_reference_that_differs_only_in_case_is_the_same_product(self, client):
        plan(client, text=BLOCK, apply=True)
        result = plan(client, text="Model Reference\tManufacturer\nsys-vsr-500\tSystemair\n")
        assert result["rows"][0]["action"] == "unchanged"

    def test_a_row_that_changes_nothing_is_reported_as_such(self, client):
        plan(client, text=BLOCK, apply=True)
        result = plan(client, text=BLOCK)
        assert result["counts"]["unchanged"] == 2
        assert result["can_apply"] is False
        assert result["destructive"] is False

    def test_updates_can_be_turned_off_entirely(self, client):
        plan(client, text=BLOCK, apply=True)
        result = plan(client, text=(
            "Model Reference\tManufacturer\tLength (mm)\nSYS-VSR-500\tSystemair\t9999\n"
        ), update_existing=False)
        assert result["rows"][0]["action"] == "skip"
        assert "already in the library" in result["rows"][0]["reason"]

    def test_an_import_that_would_change_values_says_it_is_destructive(self, client):
        plan(client, text=BLOCK, apply=True)
        result = plan(client, text=(
            "Model Reference\tLength (mm)\nSYS-VSR-500\t1250\n"
        ))
        assert result["destructive"] is True

    def test_a_blank_cell_never_deletes_what_is_already_known(self, client):
        plan(client, text=BLOCK, apply=True)
        # A supplier's sheet is usually partial. A blank Length means "not
        # stated here", not "forget the 1200 you already have".
        plan(client, text=(
            "Model Reference\tManufacturer\tLength (mm)\nSYS-VSR-500\tSystemair AB\t\n"
        ), apply=True)
        entry = library(client)["SYS-VSR-500"]["values"]
        assert entry["Length (mm)"] == 1200
        assert entry["Manufacturer"] == "Systemair AB"


class TestApplying:
    def test_applying_creates_the_products(self, client):
        result = plan(client, text=BLOCK, apply=True)
        assert result["applied"] == 2
        entries = library(client)
        assert set(entries) == {"SYS-VSR-500", "SYS-VSR-700"}
        assert entries["SYS-VSR-500"]["values"]["Manufacturer"] == "Systemair"

    def test_numbers_arrive_as_numbers(self, client):
        plan(client, text=BLOCK, apply=True)
        assert library(client)["SYS-VSR-500"]["values"]["Length (mm)"] == 1200

    def test_applying_updates_only_what_changed(self, client):
        plan(client, text=BLOCK, apply=True)
        result = plan(client, text=(
            "Model Reference\tManufacturer\tLength (mm)\n"
            "SYS-VSR-500\tSystemair\t1250\n"
            "SYS-VSR-700\tSystemair\t1400\n"
        ), apply=True)
        assert result["applied"] == 1, "the unchanged row is not rewritten"
        assert library(client)["SYS-VSR-500"]["values"]["Length (mm)"] == 1250

    def test_an_import_with_nothing_to_do_is_refused_rather_than_silently_passing(self, client):
        plan(client, text=BLOCK, apply=True)
        response = client.post("/api/library/import", json={
            "type_code": "MVHR", "text": BLOCK, "apply": True,
        })
        assert response.status_code == 400
        assert "nothing to import" in response.json()["detail"]

    def test_imported_products_are_flagged_for_review_like_any_other(self, client):
        plan(client, text=BLOCK, apply=True)
        queue = client.get("/api/library/review/queue").json()
        assert {e["model_reference"] for e in queue} >= {"SYS-VSR-500", "SYS-VSR-700"}

    def test_an_import_is_recorded_in_the_change_log(self, client):
        plan(client, text=BLOCK, apply=True)
        entries = client.get("/api/impact?area=library").json()["entries"]
        assert any("SYS-VSR-500" in e["summary"] for e in entries)

    def test_the_preview_and_the_apply_agree(self, client):
        preview = plan(client, text=BLOCK)
        applied = plan(client, text=BLOCK, apply=True)
        assert preview["counts"]["create"] == applied["applied"]
        assert [r["action"] for r in preview["rows"]] == [r["action"] for r in applied["rows"]]


class TestGridEntry:
    def test_several_products_save_in_one_go(self, client):
        response = client.post("/api/library/bulk", json={
            "type_code": "MVHR",
            "rows": [
                {"type_code": "MVHR", "model_reference": "RAD-001",
                 "values": {"Manufacturer": "Merriott"}},
                {"type_code": "MVHR", "model_reference": "RAD-002",
                 "values": {"Manufacturer": "Merriott"}},
            ],
        })
        assert response.status_code == 201
        assert len(response.json()) == 2
        assert set(library(client)) == {"RAD-001", "RAD-002"}

    def test_a_row_with_no_reference_refuses_the_whole_batch(self, client):
        response = client.post("/api/library/bulk", json={
            "type_code": "MVHR",
            "rows": [
                {"type_code": "MVHR", "model_reference": "RAD-001", "values": {}},
                {"type_code": "MVHR", "model_reference": "  ", "values": {"Manufacturer": "X"}},
            ],
        })
        assert response.status_code == 400
        assert "row(s) 2" in response.json()["detail"]
        assert library(client) == {}, "nothing is saved when the batch is refused"


class TestTheLibraryWorkbook:
    """Out as a spreadsheet, filled in, and back again.

    The paste importer asks somebody to know which column is which before they
    have seen the columns. This is the route that does not: hand them the file
    with the headings on it. What comes back has to go through the same planner
    a paste does, or the two would drift apart in what they accept.
    """

    @pytest.fixture()
    def client(self, tmp_path, monkeypatch):
        from fastapi.testclient import TestClient

        monkeypatch.setenv("SCHEDUL_DATABASE_URL", f"sqlite:///{tmp_path / 'wb.db'}")
        import schedul.db.session as session_module

        session_module.SessionLocal = None
        session_module.init_db(f"sqlite:///{tmp_path / 'wb.db'}")
        from schedul.api.main import app

        return TestClient(app)

    @staticmethod
    def _book(content):
        import io

        from openpyxl import load_workbook

        return load_workbook(io.BytesIO(content))

    @staticmethod
    def _post(client, content, **data):
        return client.post(
            "/api/library/workbook/import",
            files={"file": ("library.xlsx", content, "application/octet-stream")},
            data={"apply": "false", **data},
        )

    def test_every_type_gets_a_sheet_named_after_its_code(self, client):
        book = self._book(client.get("/api/library/workbook.xlsx").content)
        codes = {t["code"] for t in client.get("/api/catalogue").json()}
        assert codes <= set(book.sheetnames)

    def test_the_headings_are_the_lookup_key_then_the_library_fields(self, client):
        book = self._book(client.get("/api/library/workbook.xlsx?code=MVHR").content)
        sheet = book["MVHR"]
        headings = [sheet.cell(1, c).value for c in range(1, 30) if sheet.cell(1, c).value]
        assert headings[0] == "Model Reference"
        assert "Manufacturer" in headings
        assert "Unit Reference" not in headings, (
            "an input column differs per unit, so it would be a stale copy in the library"
        )

    def test_the_blank_template_carries_an_example_that_imports_as_nothing(self, client):
        template = client.get("/api/library/workbook.xlsx?code=MVHR&data=false").content
        plan = self._post(client, template).json()
        assert plan["counts"]["create"] == 0
        assert not plan["can_apply"], (
            "importing the template unedited must add nothing, example row included"
        )

    def test_an_export_comes_back_as_no_change_at_all(self, client):
        client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-VSR-500",
            "values": {"Manufacturer": "Systemair", "Width (mm)": 900},
        })
        exported = client.get("/api/library/workbook.xlsx?code=MVHR").content
        plan = self._post(client, exported).json()
        assert plan["counts"]["unchanged"] == 1
        assert plan["counts"]["create"] == 0 and plan["counts"]["update"] == 0

    def test_a_corrected_workbook_updates_and_adds(self, client):
        import io

        client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-VSR-500",
            "values": {"Manufacturer": "Systemair", "Width (mm)": 900},
        })
        book = self._book(client.get("/api/library/workbook.xlsx?code=MVHR").content)
        sheet = book["MVHR"]
        width = next(c for c in range(1, 30) if sheet.cell(1, c).value == "Width (mm)")
        sheet.cell(2, width, 950)
        sheet.cell(3, 1, "SYS-VSR-700")
        sheet.cell(3, width, 1100)
        buffer = io.BytesIO()
        book.save(buffer)

        plan = self._post(client, buffer.getvalue()).json()
        assert plan["counts"] == {"create": 1, "update": 1, "unchanged": 0, "skip": 0}
        assert plan["destructive"], "changing a stored value has to be confirmed"

        applied = self._post(client, buffer.getvalue(), apply="true").json()
        assert applied["applied"] == 2
        entries = {
            e["model_reference"]: e["values"] for e in client.get("/api/library/MVHR").json()
        }
        assert entries["SYS-VSR-500"]["Width (mm)"] == 950
        assert entries["SYS-VSR-700"]["Width (mm)"] == 1100

    def test_a_blank_cell_still_means_not_stated(self, client):
        import io

        client.post("/api/library", json={
            "type_code": "MVHR", "model_reference": "SYS-VSR-500",
            "values": {"Manufacturer": "Systemair", "Width (mm)": 900},
        })
        book = self._book(client.get("/api/library/workbook.xlsx?code=MVHR").content)
        sheet = book["MVHR"]
        width = next(c for c in range(1, 30) if sheet.cell(1, c).value == "Width (mm)")
        sheet.cell(2, width, None)
        buffer = io.BytesIO()
        book.save(buffer)

        self._post(client, buffer.getvalue(), apply="true")
        entry = client.get("/api/library/MVHR").json()[0]
        assert entry["values"]["Width (mm)"] == 900, (
            "an emptied cell means 'not stated here', never 'delete what you know'"
        )

    def test_the_same_reference_twice_is_one_product(self, client):
        import io

        book = self._book(
            client.get("/api/library/workbook.xlsx?code=MVHR&data=false").content
        )
        sheet = book["MVHR"]
        for row in (2, 3):
            sheet.cell(row, 1, "SYS-VSR-500")
            sheet.cell(row, 2, "Systemair")
        buffer = io.BytesIO()
        book.save(buffer)

        plan = self._post(client, buffer.getvalue()).json()
        actions = [r["action"] for r in plan["sheets"][0]["rows"]]
        assert actions.count("create") == 1
        assert "skip" in actions

    def test_a_sheet_naming_no_type_is_reported_not_guessed_at(self, client):
        import io

        book = self._book(
            client.get("/api/library/workbook.xlsx?code=MVHR&data=false").content
        )
        book["MVHR"].title = "Radiators maybe"
        buffer = io.BytesIO()
        book.save(buffer)

        plan = self._post(client, buffer.getvalue()).json()
        assert plan["sheets"][0]["recognised"] is False
        assert "no schedule type" in plan["sheets"][0]["message"]

    def test_something_that_is_not_a_workbook_is_refused_plainly(self, client):
        response = client.post(
            "/api/library/workbook/import",
            files={"file": ("notes.txt", b"hello", "text/plain")},
            data={"apply": "false"},
        )
        assert response.status_code == 400
        assert "Excel workbook" in response.text


class TestTheScheduleWorkbook:
    """A schedule's typed columns, out as a spreadsheet and back.

    The deliverable export is a different file with a different job: a cover, a
    revision page and every calculated column. This is the working one, and the
    only columns in it are the ones somebody types.
    """

    @pytest.fixture()
    def client(self, tmp_path, monkeypatch):
        from fastapi.testclient import TestClient

        monkeypatch.setenv("SCHEDUL_DATABASE_URL", f"sqlite:///{tmp_path / 'sw.db'}")
        import schedul.db.session as session_module

        session_module.SessionLocal = None
        session_module.init_db(f"sqlite:///{tmp_path / 'sw.db'}")
        from schedul.api.main import app

        return TestClient(app)

    @pytest.fixture()
    def schedule(self, client):
        project = client.post("/api/projects", json={"number": "CM1"}).json()
        building = project["buildings"][0]["id"]
        made = client.post(
            f"/api/projects/{project['id']}/buildings/{building}/schedules",
            json={"code": "MVHR"},
        ).json()
        return made["buildings"][0]["schedules"][0]["id"]

    @staticmethod
    def _book(content):
        import io

        from openpyxl import load_workbook

        return load_workbook(io.BytesIO(content))

    def _post(self, client, schedule, content, **data):
        return client.post(
            f"/api/schedules/{schedule}/rows/workbook",
            files={"file": ("rows.xlsx", content, "application/octet-stream")},
            data={"mode": "append", "apply": "false", **data},
        )

    def test_only_the_columns_somebody_types_are_in_it(self, client, schedule):
        sheet = self._book(
            client.get(f"/api/schedules/{schedule}/rows.xlsx").content
        ).active
        headings = [sheet.cell(1, c).value for c in range(1, 30) if sheet.cell(1, c).value]
        assert "Unit Reference" in headings
        assert "Model Reference" in headings
        assert "Manufacturer" not in headings, "a product column is looked up, not typed"
        assert "Total Airflow (l/s)" not in headings, "a calculated column is worked out"

    def test_the_blank_one_has_the_same_headings_and_no_rows(self, client, schedule):
        filled = self._book(client.get(f"/api/schedules/{schedule}/rows.xlsx").content).active
        blank = self._book(
            client.get(f"/api/schedules/{schedule}/rows.xlsx?filled=false").content
        ).active
        heads = lambda s: [s.cell(1, c).value for c in range(1, 30)]
        assert heads(filled) == heads(blank)
        assert blank.cell(2, 1).value is None

    def test_what_is_on_the_schedule_comes_out_in_it(self, client, schedule):
        client.post(f"/api/schedules/{schedule}/rows", json={
            "values": {"Unit Reference": "MVHR-001", "Location": "Cupboard"},
        })
        sheet = self._book(
            client.get(f"/api/schedules/{schedule}/rows.xlsx").content
        ).active
        assert sheet.cell(2, 1).value == "MVHR-001"
        assert sheet.cell(2, 2).value == "Cupboard"

    def test_a_filled_in_workbook_comes_back(self, client, schedule):
        import io

        book = self._book(
            client.get(f"/api/schedules/{schedule}/rows.xlsx?filled=false").content
        )
        sheet = book.active
        sheet.cell(2, 1, "MVHR-001")
        sheet.cell(2, 2, "Cupboard")
        sheet.cell(3, 1, "MVHR-002")
        buffer = io.BytesIO()
        book.save(buffer)

        plan = self._post(client, schedule, buffer.getvalue()).json()
        assert plan["detected_rows"] == 2
        assert plan["header_detected"], "a workbook's first line is a heading by construction"

        applied = self._post(client, schedule, buffer.getvalue(), apply="true").json()
        assert applied["applied"] == 2
        grid = client.get(f"/api/schedules/{schedule}").json()
        assert [r["values"].get("Unit Reference") for r in grid["rows"]] == [
            "MVHR-001", "MVHR-002"
        ]

    def test_a_gap_in_the_middle_does_not_end_the_import(self, client, schedule):
        import io

        book = self._book(
            client.get(f"/api/schedules/{schedule}/rows.xlsx?filled=false").content
        )
        sheet = book.active
        sheet.cell(2, 1, "MVHR-001")
        sheet.cell(4, 1, "MVHR-002")   # row 3 left blank
        buffer = io.BytesIO()
        book.save(buffer)

        applied = self._post(client, schedule, buffer.getvalue(), apply="true").json()
        assert applied["applied"] == 2, (
            "a schedule genuinely has empty rows; stopping at the first would read "
            "back only the part above it"
        )

    def test_replacing_is_still_refused_without_a_confirmation(self, client, schedule):
        import io

        client.post(f"/api/schedules/{schedule}/rows", json={
            "values": {"Unit Reference": "ALREADY HERE"},
        })
        book = self._book(
            client.get(f"/api/schedules/{schedule}/rows.xlsx?filled=false").content
        )
        book.active.cell(2, 1, "MVHR-001")
        buffer = io.BytesIO()
        book.save(buffer)

        response = self._post(
            client, schedule, buffer.getvalue(), mode="replace", apply="true"
        )
        assert response.status_code == 409
        grid = client.get(f"/api/schedules/{schedule}").json()
        assert grid["rows"][0]["values"]["Unit Reference"] == "ALREADY HERE"

    def test_something_that_is_not_a_workbook_is_refused_plainly(self, client, schedule):
        response = client.post(
            f"/api/schedules/{schedule}/rows/workbook",
            files={"file": ("notes.txt", b"hello", "text/plain")},
            data={"apply": "false"},
        )
        assert response.status_code == 400
        assert "Excel workbook" in response.text
