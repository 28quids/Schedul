"""The equipment library as a workbook: out, filled in, and back again.

Getting a manufacturer's range into the library by pasting a block into a
textarea works, but it asks somebody to know which column is which before they
have seen the columns. The obvious way round is the one everybody already has:
hand them the spreadsheet with the headings on it, let them fill it in where
they are comfortable, and read it back.

Three files, all the same shape, so the one that comes back is always something
this can read:

- a **blank template** for one type, headings and an example row;
- an **export of one type**, the same headings with what is already there;
- an **export of the whole library**, one sheet per type.

The last is the mass reimport: a practice can take everything out, correct it in
Excel where correcting a hundred rows is a drag of the fill handle, and bring it
back. Nothing about the read path is new -- a sheet becomes tab-separated text
and goes through ``services.importing`` like a paste, so the duplicate handling,
the "a blank cell means not stated" rule and the plan-before-it-happens
guarantee are the same ones the paste importer already has.

The lookup key is always column A and always ``Model Reference``, exactly as v1's
library sheets were.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..core.catalogue import MODEL_REFERENCE, ScheduleType

__all__ = [
    "sheet_name_for",
    "library_columns",
    "render_library_workbook",
    "read_library_workbook",
]

_HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
_HDR_FONT = Font(name="Arial", size=10, bold=True)
_BODY_FONT = Font(name="Arial", size=10)

#: Excel refuses these in a sheet name, and truncates past 31 characters.
_ILLEGAL = re.compile(r"[\[\]:*?/\\]")


def sheet_name_for(code: str) -> str:
    """One type's sheet name. The code, because the code is the key."""
    return _ILLEGAL.sub("-", str(code).strip().upper())[:31] or "SHEET"


def library_columns(schedule_type: ScheduleType) -> list[str]:
    """The headings for one type: the lookup key, then its library fields.

    Input and derived columns are deliberately absent. An input differs per unit
    and a derived one is calculated, so either would be a column of stale copies
    the moment somebody filled it in.
    """
    return [MODEL_REFERENCE, *(c.legacy_name for c in schedule_type.library)]


def render_library_workbook(
    types: Sequence[ScheduleType],
    products: dict[str, Sequence[dict[str, Any]]],
    out_path: str | Path,
    *,
    include_examples: bool = True,
) -> Path:
    """Write one workbook covering ``types``, a sheet each.

    ``products`` maps a type code to the entries to write; a code with nothing
    against it comes out as a blank sheet with its headings, which is exactly
    what a template is. There is no separate template renderer, because two
    renderers would eventually disagree about the headings and the file that
    came back would stop matching the file that went out.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for schedule_type in types:
        columns = library_columns(schedule_type)
        ws = wb.create_sheet(sheet_name_for(schedule_type.code))

        for i, heading in enumerate(columns, start=1):
            cell = ws.cell(1, i, heading)
            cell.font = _HDR_FONT
            cell.fill = _HDR_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = max(
                14, min(len(heading) + 4, 40)
            )
        ws.column_dimensions["A"].width = 26
        ws.row_dimensions[1].height = 30

        rows = list(products.get(schedule_type.code, []))
        for r, product in enumerate(rows, start=2):
            for i, key in enumerate(columns, start=1):
                ws.cell(r, i, product.get(key)).font = _BODY_FONT

        if not rows and include_examples:
            # One example row, in the same shape, so nobody has to guess whether
            # a dimension goes in as '900' or '900mm'. It is skipped on import:
            # its reference is the type's own EXAMPLE key, which is never real.
            ws.cell(2, 1, f"{schedule_type.code}-EXAMPLE-01").font = _BODY_FONT
            for i, column in enumerate(schedule_type.library, start=2):
                ws.cell(2, i, column.example).font = _BODY_FONT

        # The instructions go on the heading cell rather than into a cell of
        # their own. A sentence written under the data is a row as far as any
        # reader is concerned, and 'Type over the example row' arriving in the
        # library as a product reference is a nasty way to discover that.
        ws["A1"].comment = Comment(
            "Column A is the key: two rows with the same reference are the same "
            "product, and importing one that is already here updates it rather "
            "than adding a second. A blank cell means 'not stated', so it leaves "
            "what is already stored alone.",
            "Schedul",
            width=320,
            height=120,
        )

        # A real table, so filling one down carries the formatting with it and
        # so the range grows when somebody adds rows at the bottom.
        last = max(2, len(rows) + 1)
        table = Table(
            displayName=f"Library_{sheet_name_for(schedule_type.code).replace('-', '_')}",
            ref=f"A1:{get_column_letter(len(columns))}{last}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)
        ws.freeze_panes = "B2"

    if not wb.worksheets:  # an organisation with no types at all
        wb.create_sheet("Library")

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def read_library_workbook(source: Any) -> list[dict[str, Any]]:
    """Read a filled-in workbook back as one block of text per sheet.

    Returns ``[{"sheet": ..., "code": ..., "text": ...}]``. The text is
    tab-separated with the heading row first, which is what the paste importer
    already takes -- so a workbook import and a pasted block go down the same
    path and cannot drift apart in what they accept or how they report it.

    Values are stringified here rather than in the planner: openpyxl gives back
    a float for every number, and ``900.0`` written into the library as text
    would read as a dimension nobody typed.
    """
    wb = load_workbook(source, data_only=True, read_only=True)
    blocks: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [_text(v) for v in (row or ())]
            while cells and cells[-1] == "":
                cells.pop()
            if not cells or not any(cells):
                # A blank row ends the block, exactly as Excel's own current
                # region does. The template puts its instructions under one, and
                # a sentence of guidance imported as a product reference is a
                # nasty way to find that out.
                if lines:
                    break
                continue
            lines.append("\t".join(cells))
        if len(lines) < 2:
            continue  # headings with nothing under them is not an import
        blocks.append(
            {
                "sheet": ws.title,
                "code": _code_from_sheet(ws.title),
                "text": "\n".join(lines),
            }
        )
    wb.close()
    return blocks


def _code_from_sheet(title: str) -> str:
    """The type code a sheet is for.

    The exporter names a sheet after the code, but somebody who has renamed
    'MVHR' to 'MVHR units' still means MVHR, so the first word wins.
    """
    text = str(title or "").strip()
    for separator in (" - ", " ", "-"):
        if separator in text:
            text = text.split(separator)[0]
            break
    return text.strip().upper()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
