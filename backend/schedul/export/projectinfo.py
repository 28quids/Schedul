"""MAINPROJECTINFO.xlsx -- the project's master setup and read document.

Kept broadly as v1 had it: a ``Setup`` sheet of key/value project fields and a
``ScheduleList`` sheet. The layout is a contract, so it does not change shape.

Two differences:

- ``ScheduleList`` gains a leading ``Building`` column, so one file covers the
  whole project across every block (SPEC.md 6.2). One MAINPROJECTINFO per
  project, not per building.
- Revision, IssueDate and Status are **written as values**. Under v1 they were
  blank until someone refreshed a Power Query scrape of every workbook's
  Metadata sheet. The manager knows them, so it fills them in.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy.orm import Session

from ..core.naming import NamingError
from ..core.revisions import current as current_revision

__all__ = ["render_project_info"]

_HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
_THIN = Side(style="thin", color="808080")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def render_project_info(
    session: Session, project: Any, org: Any, out_path: str | Path
) -> Path:
    """Write one project's MAINPROJECTINFO workbook."""
    from ..services import projects as svc
    from ..services.converters import revisions_of

    house = svc.house_standard_for(session, org.id)
    scheme = svc.naming_scheme_for(session, org.id)

    wb = Workbook()
    setup = wb.active
    setup.title = "Setup"
    setup.column_dimensions["A"].width = 30
    setup.column_dimensions["B"].width = 60

    # RefreshProjectData read this sheet by key/value lookup, so the shape is a
    # contract even though the macro that consumed it is gone: anyone's existing
    # tooling that reads it keeps working.
    for i, (key, value) in enumerate(project.project_fields.items(), start=1):
        setup.cell(i, 1, key).font = Font(name="Arial", size=10, bold=True)
        setup.cell(i, 2, value).font = Font(name="Arial", size=10)

    listing = wb.create_sheet("ScheduleList")
    headers = [
        "Building", "DocumentNumber", "ScheduleName", "Revision",
        "IssueDate", "Status", "StatusDescription", "FileName",
    ]
    for i, header in enumerate(headers, start=1):
        c = listing.cell(1, i, header)
        c.font = Font(name="Arial", size=10, bold=True)
        c.fill = _HDR_FILL
        c.border = _BOX
    for col, width in zip("ABCDEFGH", [18, 58, 46, 10, 12, 10, 30, 70]):
        listing.column_dimensions[col].width = width

    row = 2
    for building in svc.buildings_of(session, project):
        for schedule in svc.live_schedules(session, building):
            try:
                docnum = svc.document_number_for(schedule, scheme)
                filename = svc.filename_for(schedule, scheme)
            except NamingError:
                docnum, filename = schedule.docnum, ""

            latest = current_revision(revisions_of(schedule))
            status_code, status_desc = "", ""
            if latest and latest.status:
                text = str(latest.status)
                if " - " in text:
                    status_code, status_desc = text.split(" - ", 1)
                else:
                    status_code = text
                    status_desc = house.status_description(text)

            values = [
                building.label,
                docnum,
                schedule.schedule_type.title if schedule.schedule_type else schedule.code,
                latest.code if latest else "",
                latest.date if latest else None,
                status_code,
                status_desc,
                filename,
            ]
            for i, value in enumerate(values, start=1):
                cell = listing.cell(row, i, value)
                cell.font = Font(name="Arial", size=10)
                cell.border = _BOX
            listing.cell(row, 5).number_format = "DD/MM/YYYY"
            row += 1

    listing.freeze_panes = "A2"
    listing.auto_filter.ref = f"A1:H{max(row - 1, 1)}"

    note = listing.cell(
        row + 1, 1,
        "Exported from Schedul. Revision, issue date and status are written as "
        "values from the record, so nothing needs refreshing.",
    )
    note.font = Font(name="Arial", size=9, italic=True, color="595959")

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
