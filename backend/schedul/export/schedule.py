"""Rendering one schedule workbook.

The sheet-construction code is vendored from v1's ``build_schedule_file`` and
kept deliberately close to it: roughly 350 lines of Excel minutiae that is
correct and hand-checked against the real house files. A4 portrait covers and
landscape schedules, the Verdana 30pt title block, the two-row header with units
split onto row 5, print titles ``$1:$5``, ``fitToWidth=1 / fitToHeight=0``, the
INDEX/MATCH product lookups and the blue/green/black colour contract all survive
unchanged. SPEC.md 6.2: refactor its interface, not its body.

What changes:

1. **The document number is a parameter**, not a loop index. That is the whole
   of SPEC.md fact 1.
2. **Columns arrive in the new three-kind shape**, adapted back to the legacy
   parallel lists internally so the body keeps working.
3. **The revision summary is driven by a sort key**, not by "last non-empty
   row", fixing the bug both v1 implementations have (SPEC.md 6.1).
4. **The summary block is built from an ordered list of (label, source) pairs**
   so row numbers fall out rather than being hardcoded, which is what makes
   adding the Building row cost one list entry (SPEC.md 4.6).
5. **Notes are two-source**: project notes then type notes, numbered
   continuously (SPEC.md 4.7).
6. **Live data is written in**, because the database is the record and the
   workbook is an export of it.

The exported workbook is macro-free and self-contained: the equipment library
for this type is embedded, so the file calculates on any machine with no add-in,
no shared path and no external links. The ``path_*`` Config keys from v1 are
gone with the macros that read them.
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..core.branding import Branding
from ..core.catalogue import MODEL_REFERENCE, ScheduleType
from ..core.formula import CONSTANTS, FormulaError, to_excel
from ..core.house import HouseStandard
from ..core.revisions import PRELIMINARY_BASE, PUBLISHED_BASE, Revision, rank
from ..core.units import split_unit

__all__ = ["ScheduleContent", "render_schedule"]

A4 = 9

FILL_IN = PatternFill("solid", fgColor="FFFFCC")
FILL_HDR = PatternFill("solid", fgColor="D9D9D9")
FILL_GRP_IN = PatternFill("solid", fgColor="DCE6F1")
FILL_GRP_LIB = PatternFill("solid", fgColor="E2EFDA")
FILL_GRP_CALC = PatternFill("solid", fgColor="FCE4D6")

THIN = Side(style="thin", color="808080")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
TOPL = Alignment(horizontal="left", vertical="top", wrap_text=True)


class _Style:
    """Fonts derived from the house style. Vendored from v1."""

    def __init__(self, hs: dict[str, Any]) -> None:
        cf, sf = hs["cover_font"], hs["schedule_font"]
        self.grey = hs["title_grey"][2:]
        self.blue = hs["title_blue"][2:]
        self.f_big_grey = Font(name=cf, size=hs["title_size"], bold=True, color=self.grey)
        self.f_big_blue = Font(name=cf, size=hs["title_size"], bold=True, color=self.blue)
        self.f_cov_lbl = Font(name=cf, size=hs["cover_body_size"], bold=False)
        self.f_cov_val = Font(name=cf, size=hs["cover_body_size"], bold=True)
        self.f_cov_sm = Font(name=cf, size=9, bold=True)
        self.f_rev_hdr = Font(name=cf, size=9, bold=True)
        self.f_rev_bod = Font(name=cf, size=9)
        self.f_rev_in = Font(name=cf, size=9, color="0000FF")
        self.f_title = Font(name=sf, size=12, bold=True)
        self.f_note = Font(name=sf, size=8)
        self.f_hdr = Font(name=sf, size=8, bold=True)
        self.f_unit = Font(name=sf, size=8, italic=True, color="595959")
        self.f_in = Font(name=sf, size=8, color="0000FF")
        self.f_pull = Font(name=sf, size=8, color="008000")
        self.f_calc = Font(name=sf, size=8, color="000000")
        self.f_sm = Font(name=sf, size=8)


#: Roughly how many characters of an 11pt font fit in one unit of column width.
#: Excel measures a column in characters of the standard font, so a title at
#: another size scales inversely.
_CHARS_PER_UNIT = 11.0


def _fit_title(text: str, chars_available: float, wanted: int, max_lines: int = 2):
    """A point size and line count that fit ``text`` across the merged title.

    The house cover sets the project name and the schedule title at 30pt across
    seven columns. That is fine for 'Fan Coil Unit Schedule' and runs clean off
    the page for 'Mechanical Ventilation with Heat Recovery Unit Schedule',
    which is what was happening: the text was clipped at the right margin and
    overlapped the line below it.

    So it wraps first, up to two lines, and only shrinks when even that will not
    fit. The practice's chosen size is the maximum rather than the mandate,
    because a title that has been silently cut in half is worse than one set a
    few points smaller.
    """
    import math

    text = str(text or "")
    size = max(int(wanted), 8)
    while size > 10:
        per_line = max(1, int(chars_available * _CHARS_PER_UNIT / size))
        lines = max(1, math.ceil(len(text) / per_line))
        if lines <= max_lines:
            return size, lines
        size -= 1
    per_line = max(1, int(chars_available * _CHARS_PER_UNIT / size))
    return size, max(1, math.ceil(len(text) / per_line))


def _page(ws, landscape: bool) -> None:
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = A4
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False


class ScheduleContent:
    """Everything the renderer needs, gathered by the caller.

    A plain container rather than database rows, so the renderer stays testable
    without a session and the export layer never reaches into the ORM.
    """

    def __init__(
        self,
        *,
        schedule_type: ScheduleType,
        house: HouseStandard,
        project_fields: dict[str, str],
        design_constants: dict[str, float],
        docnum: str,
        building_ref: str = "",
        building_name: str = "",
        rows: Sequence[dict[str, Any]] = (),
        overrides: Sequence[dict[str, Any]] = (),
        computed: Sequence[dict[str, Any]] = (),
        frozen: bool = False,
        revisions: Sequence[Revision] = (),
        products: Sequence[dict[str, Any]] = (),
        doc_type: str = "SC",
        classification: str = "",
        theme: str = "xlsx",
        notes: Sequence[str] | None = None,
    ) -> None:
        self.schedule_type = schedule_type
        self.house = house
        self.project_fields = project_fields
        self.design_constants = design_constants
        self.docnum = docnum
        self.building_ref = building_ref
        self.building_name = building_name
        self.rows = list(rows)
        self.overrides = list(overrides)
        #: Library and derived values as they were when a revision was issued.
        self.computed = list(computed)
        #: A frozen export writes those values literally instead of formulas.
        #: A formula would recompute against today's library, which is exactly
        #: what an issued document must not do.
        self.frozen = frozen
        #: 'xlsx' keeps the editing colours so the file stays workable; 'pdf'
        #: uses the issue theme, because an issued document should not look like
        #: somebody's editing screen.
        self.theme = theme
        self.revisions = list(revisions)
        self.products = list(products)
        self.doc_type = doc_type
        self.classification = classification
        #: The resolved notes, already layered by core.notes. None falls back to
        #: the organisation's plus the type's, which is what they resolve to for
        #: a schedule that has not diverged.
        self.notes = list(notes) if notes is not None else None

    @property
    def issue_theme(self) -> bool:
        """Whether to render plainly, as an issued document rather than a form.

        'issue' and 'pdf' both mean a document that is being sent to somebody.
        'editor' keeps the yellow input fill and the blue/green/black contract,
        which are editing aids and belong on a file somebody is still filling in.
        """
        return self.theme in ("pdf", "issue")

    def overrides_for(self, index: int) -> dict[str, Any]:
        return self.overrides[index] if index < len(self.overrides) else {}

    def computed_for(self, index: int) -> dict[str, Any]:
        return self.computed[index] if index < len(self.computed) else {}

    @property
    def branding(self) -> Branding:
        """The practice's document appearance, as the renderer's instructions."""
        return Branding.from_dict(self.house.branding)

    @property
    def house_style(self) -> dict[str, Any]:
        """The house style with branding's fonts and colours applied over it.

        Branding is an overlay rather than a second copy: the house style also
        holds row counts and body sizes that have nothing to do with branding,
        and two places holding the cover font would eventually disagree.
        """
        return {**self.house.house_style, **self.branding.house_style_overrides()}

    @property
    def building_label(self) -> str:
        if self.building_ref and self.building_name:
            return f"{self.building_ref} - {self.building_name}"
        return self.building_ref or self.building_name


def _notes_block(content: ScheduleContent) -> str:
    """Project notes then type notes, numbered continuously across both.

    SPEC.md 4.7 and 1a.1: the real house file's A2 is equipment-specific
    ("radiant panels are to be sized with a 55degC flow"), while v1 put the same
    project-level block on every schedule. Both belong, in that order.

    Which notes those are is resolved by ``core.notes`` before it gets here --
    organisation, then project, then type, unless the schedule has taken them
    over -- so the workbook prints exactly what the editor showed.
    """
    combined = (
        content.notes
        if content.notes is not None
        else [*content.house.general_notes, *content.schedule_type.notes]
    )
    if not combined:
        return ""
    numbered = "\n".join(f"[{i}] {n}" for i, n in enumerate(combined, start=1))
    return "General Notes:\n" + numbered


def render_schedule(content: ScheduleContent, out_path: str | Path) -> Path:
    """Write one schedule workbook and return where it landed."""
    hs = content.house_style
    st = _Style(hs)
    brand = content.branding
    stype = content.schedule_type

    # An issued document is read, not filled in. The yellow input fill and the
    # blue/green/black colour contract are editing aids; carrying them onto a
    # PDF that goes to a client makes it look like a working file.
    issue = content.issue_theme
    header_fill = PatternFill("solid", fgColor=brand.rgb("header"))
    fill_in = PatternFill("solid", fgColor="FFFFFF") if issue else FILL_IN
    grp_in = header_fill if issue else FILL_GRP_IN
    grp_lib = header_fill if issue else FILL_GRP_LIB
    grp_calc = header_fill if issue else FILL_GRP_CALC
    font_in = st.f_sm if issue else st.f_in
    font_pull = st.f_sm if issue else st.f_pull

    inputs = stype.inputs
    library = stype.library
    derived = stype.derived

    # A working file carries the house standard's spare rows, because somebody is
    # about to fill them in. An issued one carries what it says: forty ruled
    # empty rows under six units reads as an unfinished document, and takes
    # pages of a PDF to say nothing.
    if issue:
        data_rows = max(len(content.rows), 1)
        rev_rows = max(len(content.revisions), 1)
    else:
        data_rows = max(int(hs.get("data_rows", 40)), len(content.rows) + 5)
        rev_rows = max(int(hs.get("revision_rows", 20)), len(content.revisions) + 5)

    n_inputs = len(inputs)
    mr_col = n_inputs + 1
    typ_start = mr_col + 1
    der_start = typ_start + len(library)
    n_cols = der_start + len(derived) - 1

    #: Column letter for every field a formula may reference.
    colmap: dict[str, str] = {
        c.legacy_name: get_column_letter(i) for i, c in enumerate(inputs, start=1)
    }
    colmap[MODEL_REFERENCE] = get_column_letter(mr_col)
    for i, c in enumerate(library, start=typ_start):
        colmap[c.legacy_name] = get_column_letter(i)
    for i, c in enumerate(derived, start=der_start):
        colmap[c.legacy_name] = get_column_letter(i)

    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------- Config --
    cf = wb.create_sheet("Config")
    cf.column_dimensions["A"].width = 40
    cf.column_dimensions["B"].width = 70
    cf["A1"], cf["B1"] = "Key", "Value"
    for c in ("A1", "B1"):
        cf[c].font = Font(name="Arial", size=9, bold=True)

    conf_rows: dict[str, int] = {}
    row_cursor = 2

    def put(key: str, value: Any) -> None:
        nonlocal row_cursor
        cf.cell(row_cursor, 1, key).font = Font(name="Arial", size=9)
        cf.cell(row_cursor, 2, value).font = Font(name="Arial", size=9)
        conf_rows[key] = row_cursor
        row_cursor += 1

    put("EquipmentCode", stype.code)
    put("ScheduleName", stype.title)
    put("DocumentNumber", content.docnum)
    put("Building", content.building_label)
    put("BuildingRef", content.building_ref)
    for key, value in content.project_fields.items():
        put(key, value)
    for key, value in content.design_constants.items():
        put(key, value)

    for alias, full_name in CONSTANTS.items():
        if full_name in conf_rows:
            wb.defined_names.add(
                DefinedName(alias, attr_text=f"Config!$B${conf_rows[full_name]}")
            )

    cf.cell(
        row_cursor + 1, 1,
        "Exported from Schedul. The manager holds the record; this workbook is a "
        "snapshot of it. Editing here does not change the record.",
    ).font = Font(name="Arial", size=8, italic=True, color="595959")

    # -------------------------------------------------------------- Lists --
    ls = wb.create_sheet("Lists")
    ls.column_dimensions["A"].width = 40
    ls.column_dimensions["C"].width = 14
    ls["A1"], ls["C1"] = "Status", "Revision"
    for i, (code, desc) in enumerate(content.house.status_codes, start=2):
        ls.cell(i, 1, f"{code} - {desc}")
    max_rev = int(content.house.revision_codes.get("max", 20))
    revs = [f"P{n:02d}" for n in range(1, max_rev + 1)] + [
        f"C{n:02d}" for n in range(1, max(10, max_rev // 2) + 1)
    ]
    for i, rv_code in enumerate(revs, start=2):
        ls.cell(i, 3, rv_code)
    st_last = len(content.house.status_codes) + 1
    rv_last = len(revs) + 1

    # ------------------------------------------------------------ Library --
    lb = wb.create_sheet("Library")
    lib_fields = [MODEL_REFERENCE] + [c.legacy_name for c in library]
    for i, field in enumerate(lib_fields, start=1):
        name, unit = split_unit(field)
        c = lb.cell(1, i, name)
        c.font = st.f_hdr
        c.fill = FILL_HDR
        c.border = BOX
        lb.cell(2, i, unit).font = st.f_unit
        lb.column_dimensions[get_column_letter(i)].width = 18
    lb.column_dimensions["A"].width = 26

    LIB_TOP, LIB_BOT = 3, 2000
    if content.products:
        for r, product in enumerate(content.products, start=LIB_TOP):
            lb.cell(r, 1, product.get(MODEL_REFERENCE, "")).font = st.f_sm
            for j, col in enumerate(library, start=2):
                lb.cell(r, j, product.get(col.legacy_name)).font = st.f_sm
    else:
        # Keep v1's example row so the lookups demonstrate themselves on an
        # empty library rather than showing NOT FOUND everywhere.
        lb.cell(LIB_TOP, 1, f"{stype.code}-EXAMPLE-01").font = st.f_sm
        for j, col in enumerate(library, start=2):
            lb.cell(LIB_TOP, j, col.example).font = st.f_sm
    lb.freeze_panes = "B3"

    # ----------------------------------------------------------- Metadata --
    md = wb.create_sheet("Metadata")
    md.column_dimensions["A"].width = 22
    md.column_dimensions["B"].width = 60
    md["A1"], md["B1"] = "Column1", "Column2"

    # ------------------------------------------------------ Revision page --
    rv = wb.create_sheet("Revision page")
    for col, w in zip("ABCDEFG", [16, 34, 12, 13, 13, 13, 48]):
        rv.column_dimensions[col].width = w

    # The title block, sized to the width it actually has. See _fit_title.
    rv_width = sum(rv.column_dimensions[c].width for c in "ABCDEFG")
    rv_size, rv_lines = _fit_title(stype.title, rv_width, hs["title_size"])

    rv.merge_cells("A3:G3")
    rv["A3"] = f"=Config!$B${conf_rows['Project Name']}"
    rv["A3"].font = Font(
        name=hs["cover_font"], size=rv_size, bold=True, color=st.grey
    )
    rv["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rv.merge_cells("A4:G4")
    rv["A4"] = stype.title.upper()
    rv["A4"].font = Font(
        name=hs["cover_font"], size=rv_size, bold=True, color=st.blue
    )
    rv["A4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    rv.row_dimensions[3].height = max(40, rv_size * 1.45)
    rv.row_dimensions[4].height = max(40, rv_size * 1.45 * rv_lines)

    # The summary block starts at row 10 and its length is not fixed: adding a
    # field costs one entry in this list, and nothing downstream hardcodes a row.
    SUMMARY_TOP = 10
    log_top = SUMMARY_TOP + 34
    log_hdr = log_top - 1
    log_bot = log_top + rev_rows - 1

    KEY = "H"  # hidden sort-key helper column
    key_range = f"${KEY}${log_top}:${KEY}${log_bot}"

    def latest(col: str, *, blank_zero: bool = True) -> str:
        """INDEX the revision log at the row with the highest sort key.

        This is the fix. v1 used ``INDEX(range, MAX(1, COUNTA(range)))``, which
        reads the last non-empty row and so breaks on out-of-order or gapped
        entries. The real house file ranked by number after stripping ``P``,
        which sorts every C-revision below every P-revision. Ranking by the
        helper column handles both.
        """
        expr = (
            f"INDEX(${col}${log_top}:${col}${log_bot},"
            f"MATCH(MAX({key_range}),{key_range},0))"
        )
        if not blank_zero:
            return expr
        return f'IF({expr}="","",IF({expr}=0,"",{expr}))'

    status_expr = latest("B", blank_zero=False)

    # Every row the summary block can carry, keyed the way core.branding names
    # them. Which of these are written, and in what order, is the practice's
    # decision -- but the ones the cover and the Metadata sheet read by formula
    # are marked as not optional there and cannot be dropped, so a configured
    # document can never come out with a broken reference in it.
    summary_values: dict[str, tuple[Any, str | None]] = {
        "project_name": (f"=Config!$B${conf_rows['Project Name']}", None),
        "project_number": (f"=Config!$B${conf_rows['Project Number']}", None),
        "building": (f"=Config!$B${conf_rows['Building']}", None),
        "recipient": (f"=Config!$B${conf_rows['Client']}", None),
        "document_type": (content.doc_type, None),
        "revision": (f'=IFERROR({latest("A")},"")', None),
        "date": (f'=IFERROR({latest("C")},"")', "DD/MM/YYYY"),
        "prepared_by": (f'=IFERROR({latest("D")},"")', None),
        "checked_by": (f'=IFERROR({latest("E")},"")', None),
        "approved_by": (f'=IFERROR({latest("F")},"")', None),
        "document_number": (f"=Config!$B${conf_rows['DocumentNumber']}", None),
        "status": (
            f'=IFERROR(LEFT({status_expr},FIND(" -",{status_expr})-1),"")', None,
        ),
        "status_description": (
            f'=IFERROR(MID({status_expr},FIND("- ",{status_expr})+2,200),"")', None,
        ),
        "schedule_name": (stype.title, None),
        "classification": (content.classification, None),
        "bsuid": ("", None),
        "trigger_events": ("", None),
    }

    layout_fields = brand.revision_layout()
    summary_rows: dict[str, int] = {}
    for offset, field_spec in enumerate(layout_fields):
        r = SUMMARY_TOP + offset
        summary_rows[field_spec.key] = r
        value, fmt = summary_values.get(field_spec.key, ("", None))
        rv.cell(r, 1, field_spec.label).font = st.f_cov_lbl
        c = rv.cell(r, 2, value)
        c.font = st.f_cov_val
        c.alignment = LFT
        if fmt:
            c.number_format = fmt

    derived_last = SUMMARY_TOP + len(layout_fields) - 1
    rv.cell(
        derived_last + 3, 1,
        f"Rows {SUMMARY_TOP} to {derived_last} derive from the "
        f"most recent revision below, ranked by series then number so a published "
        f"C-revision outranks every preliminary one. Do not type into them.",
    ).font = Font(name=hs["cover_font"], size=8, italic=True, color="595959")

    def summary_ref(key: str) -> str:
        """A reference to one summary row, for the cover and the Metadata sheet.

        Only ever called for rows core.branding refuses to hide, so the lookup
        cannot miss -- but if a future field is made optional and something
        still reads it, this fails loudly here rather than writing #REF! into a
        document somebody issues.
        """
        row = summary_rows.get(key)
        if row is None:
            raise KeyError(
                f"the workbook reads the {key!r} row, so it cannot be hidden; "
                f"mark it as not optional in core.branding"
            )
        return f"'Revision page'!B{row}"

    headers = [
        "Revision", "Status", "Date", "Prepared by", "Checked by",
        "Approved by", "Description", "SortKey",
    ]
    for i, h in enumerate(headers, start=1):
        c = rv.cell(log_hdr, i, h)
        c.font = st.f_rev_hdr
        c.fill = FILL_HDR
        c.border = BOX
        c.alignment = CTR

    for k in range(rev_rows):
        r0 = log_top + k
        for i in range(1, 8):
            c = rv.cell(r0, i)
            c.border = BOX
            c.font = st.f_sm if issue else st.f_rev_in
            c.fill = fill_in
            c.alignment = LFT
        rv.cell(r0, 3).number_format = "DD/MM/YYYY"
        # 1000 + n for Pnn, 2000 + n for Cnn, 0 for an empty or unreadable row.
        rv.cell(
            r0, 8,
            f'=IF($A{r0}="",0,IFERROR(IF(UPPER(LEFT($A{r0},1))="C",{PUBLISHED_BASE},'
            f'{PRELIMINARY_BASE})+VALUE(MID($A{r0},2,3)),0))',
        ).font = st.f_rev_bod

    log = rank(content.revisions) if content.revisions else []
    if not log:
        log = [
            Revision(
                code="P01",
                status="S2 - Suitable for Information",
                date=_dt.date.today(),
                prepared_by=content.project_fields.get("Prepared By", ""),
                checked_by=content.project_fields.get("Checked By", ""),
                approved_by=content.project_fields.get("Approved By", ""),
                description="First issue",
            )
        ]
    for k, revision in enumerate(log):
        r0 = log_top + k
        rv.cell(r0, 1, revision.code)
        rv.cell(r0, 2, revision.status)
        rv.cell(r0, 3, revision.date)
        rv.cell(r0, 4, revision.prepared_by)
        rv.cell(r0, 5, revision.checked_by)
        rv.cell(r0, 6, revision.approved_by)
        rv.cell(r0, 7, revision.description)

    dv_s = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${st_last}", allow_blank=True)
    dv_r = DataValidation(type="list", formula1=f"=Lists!$C$2:$C${rv_last}", allow_blank=True)
    rv.add_data_validation(dv_s)
    rv.add_data_validation(dv_r)
    dv_s.add(f"B{log_top}:B{log_bot}")
    dv_r.add(f"A{log_top}:A{log_bot}")

    # A real table so the block grows when someone adds a row, and so the sort
    # key auto-fills with it, rather than being a fixed range.
    table = Table(displayName="RevisionTable", ref=f"A{log_hdr}:H{log_bot}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=False, showColumnStripes=False
    )
    rv.add_table(table)
    rv.column_dimensions[KEY].hidden = True

    _page(rv, landscape=False)
    rv.print_area = f"A1:G{log_bot}"

    # -------------------------------------------------------- Front Cover --
    fcv = wb.create_sheet("Front Cover")
    # Ten columns of 11.5 characters is about 115mm, which leaves an A4 portrait
    # page with sensible margins rather than a title block that runs off it.
    for i in range(1, 11):
        fcv.column_dimensions[get_column_letter(i)].width = 11.5

    # What each cover field says. Which of them appear, and in what order, is
    # the practice's decision -- so the rows they land on are worked out from
    # the layout rather than being numbered by hand.
    cover_values: dict[str, tuple[Any, str | None]] = {
        "recipient": (f"=Config!$B${conf_rows['Client']}", None),
        "date": (f"={summary_ref('date')}", "DD/MM/YYYY"),
        "document_number": (f"={summary_ref('document_number')}", None),
        "revision": (
            f"={summary_ref('revision')}&\"  \"&{summary_ref('status')}", None,
        ),
        # Who signed it off. The revision page already derives these from the
        # log; the cover is where a reader looks first.
        "prepared_by": (f"={summary_ref('prepared_by')}", None),
        "checked_by": (f"={summary_ref('checked_by')}", None),
        "approved_by": (f"={summary_ref('approved_by')}", None),
        "building": (f"=Config!$B${conf_rows['Building']}", None),
    }

    COVER_TOP = 11
    cover_layout = brand.cover_layout()
    for offset, field_spec in enumerate(cover_layout):
        value, fmt = cover_values.get(field_spec.key, ("", None))
        label_row = COVER_TOP + offset * 3
        fcv.cell(label_row, 1, field_spec.label).font = st.f_cov_sm
        cell = fcv.cell(label_row + 1, 1, value)
        cell.font = st.f_cov_sm
        if fmt:
            cell.number_format = fmt

    # The title block sits below the fields, wherever they end, so hiding one
    # tightens the page instead of leaving a hole in it.
    title_row = max(41, COVER_TOP + len(cover_layout) * 3 + 2)
    cover_width = sum(fcv.column_dimensions[get_column_letter(i)].width for i in range(1, 8))
    title_size, title_lines = _fit_title(stype.title, cover_width, hs["title_size"])
    wrapped = Alignment(horizontal="left", vertical="center", wrap_text=True)

    fcv.merge_cells(f"A{title_row}:G{title_row}")
    fcv[f"A{title_row}"] = f"=Config!$B${conf_rows['Project Name']}"
    fcv[f"A{title_row}"].font = Font(
        name=hs["cover_font"], size=title_size, bold=True, color=st.grey
    )
    fcv[f"A{title_row}"].alignment = wrapped
    fcv.merge_cells(f"A{title_row + 1}:G{title_row + 1}")
    fcv[f"A{title_row + 1}"] = stype.title.upper()
    fcv[f"A{title_row + 1}"].font = Font(
        name=hs["cover_font"], size=title_size, bold=True, color=st.blue
    )
    fcv[f"A{title_row + 1}"].alignment = wrapped
    # Room for the size that was actually used, and for a title that wrapped.
    fcv.row_dimensions[title_row].height = max(40, title_size * 1.45)
    fcv.row_dimensions[title_row + 1].height = max(40, title_size * 1.45 * title_lines)

    cover_bottom = title_row + 1

    # The building, under the project name (SPEC.md 4.6). One stored copy on
    # Config; this and the Revision page both read it.
    if any(f.key == "building" for f in cover_layout):
        cover_bottom += 1
        fcv.merge_cells(f"A{cover_bottom}:G{cover_bottom}")
        fcv[f"A{cover_bottom}"] = f"=Config!$B${conf_rows['Building']}"
        fcv[f"A{cover_bottom}"].font = st.f_cov_val
        fcv[f"A{cover_bottom}"].alignment = LFT

    if brand.cover_subtitle:
        cover_bottom += 1
        fcv.merge_cells(f"A{cover_bottom}:G{cover_bottom}")
        fcv[f"A{cover_bottom}"] = brand.cover_subtitle
        fcv[f"A{cover_bottom}"].font = st.f_cov_lbl
        fcv[f"A{cover_bottom}"].alignment = LFT

    if brand.cover_footer:
        cover_bottom += 2
        fcv.merge_cells(f"A{cover_bottom}:G{cover_bottom}")
        fcv[f"A{cover_bottom}"] = brand.cover_footer
        fcv[f"A{cover_bottom}"].font = Font(
            name=hs["cover_font"], size=8, color="595959"
        )
        fcv[f"A{cover_bottom}"].alignment = LFT

    _apply_branding(fcv, content)

    _page(fcv, landscape=False)
    # The print area follows the content: a fixed A1:G50 clipped a cover with a
    # footer on it and left half a page of white on one without.
    fcv.print_area = f"A1:G{max(50, cover_bottom + 2)}"

    # ---------------------------------------------------------- Schedule ---
    sc = wb.create_sheet("Schedule")
    HDR, UNIT, DAT = 4, 5, 6
    data_bot = DAT + data_rows - 1

    def measured(col, index: int, minimum: float) -> float:
        """Width that fits the header and the longest value actually present.

        The declared width is a hint from the designer, but a column of
        'ISO ePM1 60%' under a header of 'Supply Filter Grade' arrives clipped if
        nothing looks at the content. The header wraps over two lines, so it is
        allowed to count for half its length.
        """
        longest = max((len(str(v)) for v in _column_values(content, col)), default=0)
        header = max(len(w) for w in col.name.split()) if col.name else 0
        wrapped_header = max(header, len(col.name) / 2)
        return min(max(col.width * 0.8, minimum, longest + 2, wrapped_header + 2), 42)

    for i, col in enumerate(inputs, start=1):
        sc.column_dimensions[get_column_letter(i)].width = measured(col, i, 7)
    sc.column_dimensions[get_column_letter(mr_col)].width = max(
        20, min(42, max((len(str(r.get(MODEL_REFERENCE, ""))) for r in content.rows), default=0) + 2)
    )
    for i, col in enumerate(library, start=typ_start):
        sc.column_dimensions[get_column_letter(i)].width = measured(col, i, 8)
    for i, col in enumerate(derived, start=der_start):
        sc.column_dimensions[get_column_letter(i)].width = measured(col, i, 8)

    last_col = get_column_letter(n_cols)
    sc.merge_cells(f"A1:{last_col}1")
    sc["A1"] = stype.title
    sc["A1"].font = st.f_title
    sc["A1"].alignment = LFT

    sc.merge_cells(f"A2:{last_col}3")
    sc["A2"] = _notes_block(content)
    sc["A2"].font = st.f_note
    sc["A2"].alignment = TOPL
    sc.row_dimensions[2].height = 46
    sc.row_dimensions[3].height = 46

    all_columns = [*inputs, None, *library, *derived]  # None marks Model Reference
    fills = [grp_in] * mr_col + [grp_lib] * len(library) + [grp_calc] * len(derived)
    for i, col in enumerate(all_columns, start=1):
        name = MODEL_REFERENCE if col is None else col.name
        unit = "" if col is None else col.unit
        h = sc.cell(HDR, i, name)
        h.font = st.f_hdr
        h.fill = fills[i - 1]
        h.border = BOX
        h.alignment = CTR
        u = sc.cell(UNIT, i, _pretty(unit))
        u.font = st.f_unit
        u.fill = fills[i - 1]
        u.border = BOX
        u.alignment = CTR
    sc.row_dimensions[HDR].height = 46

    for i, col in enumerate(derived, start=der_start):
        if col.note:
            sc.cell(HDR, i).comment = Comment(col.note, "Schedul", width=330, height=130)

    anchor = "A"
    mrl = get_column_letter(mr_col)
    parsed: dict[str, Any] = {}

    for k in range(data_rows):
        r0 = DAT + k
        for i in range(1, n_cols + 1):
            c = sc.cell(r0, i)
            c.border = BOX
            c.alignment = LFT
            c.font = st.f_sm
        for i in range(1, mr_col + 1):
            sc.cell(r0, i).font = font_in
            sc.cell(r0, i).fill = fill_in

        for j, col in enumerate(library):
            if content.frozen:
                continue  # written literally below, from the snapshot
            lc = get_column_letter(2 + j)
            c = sc.cell(r0, typ_start + j)
            lookup = (
                f"INDEX(Library!${lc}${LIB_TOP}:${lc}${LIB_BOT},"
                f"MATCH(${mrl}{r0},Library!$A${LIB_TOP}:$A${LIB_BOT},0))"
            )
            # The inner IF matters: INDEX into an empty cell returns 0, so a
            # product with a field not yet filled in would show a hard 0 rather
            # than a blank, and 0 reads as a real duty on a schedule.
            c.value = (
                f'=IF(${mrl}{r0}="","",IFERROR(IF({lookup}="","",{lookup}),"NOT FOUND"))'
            )
            c.font = font_pull

        for j, col in enumerate(derived):
            if content.frozen:
                continue  # written literally below, from the snapshot
            c = sc.cell(r0, der_start + j)
            node = parsed.get(col.legacy_name)
            if node is None:
                try:
                    node = stype.parse_formula(col)
                except FormulaError:
                    c.value = None
                    continue
                parsed[col.legacy_name] = node
            body = to_excel(node, lambda name: f"${colmap[name]}{r0}")
            c.value = f'=IF(${anchor}{r0}="","",IFERROR({body},""))'
            c.font = st.f_calc
            c.number_format = "0.00"

    # Live data. Only what the user typed: library and derived cells already
    # carry formulas that recompute from it.
    for k, row in enumerate(content.rows):
        r0 = DAT + k
        for i, col in enumerate(inputs, start=1):
            value = row.get(col.legacy_name, row.get(col.name))
            if value not in (None, ""):
                sc.cell(r0, i, value)
        ref = row.get(MODEL_REFERENCE)
        if ref:
            sc.cell(r0, mr_col, ref)

        # A row that deliberately diverges from the library carries the literal
        # value instead of the INDEX/MATCH formula. Leaving the formula would
        # silently discard the override the moment the file is opened.
        row_overrides = content.overrides_for(k)
        for j, col in enumerate(library):
            if col.legacy_name in row_overrides:
                cell = sc.cell(r0, typ_start + j, row_overrides[col.legacy_name])
                cell.font = st.f_in if not issue else st.f_sm

        # A frozen export carries what the document said, not what a formula
        # would say today.
        if content.frozen:
            snapshot = content.computed_for(k)
            for j, col in enumerate(library):
                if col.legacy_name not in row_overrides:
                    sc.cell(r0, typ_start + j, snapshot.get(col.legacy_name)).font = font_pull
            for j, col in enumerate(derived):
                cell = sc.cell(r0, der_start + j, snapshot.get(col.legacy_name))
                cell.font = st.f_calc
                cell.number_format = "0.00"

    if not content.rows:
        for i, col in enumerate(inputs, start=1):
            sc.cell(DAT, i, col.example)
        sc.cell(DAT, mr_col, f"{stype.code}-EXAMPLE-01")

    dv_m = DataValidation(
        type="list", formula1=f"=Library!$A${LIB_TOP}:$A${LIB_BOT}", allow_blank=True
    )
    sc.add_data_validation(dv_m)
    dv_m.add(f"{mrl}{DAT}:{mrl}{data_bot}")

    _page(sc, landscape=True)
    sc.print_title_rows = f"$1:${UNIT}"
    sc.print_area = f"A1:{last_col}{data_bot}"
    sc.freeze_panes = sc.cell(DAT, 1).coordinate

    # ---------------------------------------------------- Metadata values --
    for i, (key, value) in enumerate(
        [
            ("DocumentNumber", f"={summary_ref('document_number')}"),
            ("ScheduleName", f"={summary_ref('schedule_name')}"),
            ("Revision", f"={summary_ref('revision')}"),
            ("IssueDate", f"={summary_ref('date')}"),
            ("Status", f"={summary_ref('status')}"),
            ("StatusDescription", f"={summary_ref('status_description')}"),
            ("Building", f"=Config!$B${conf_rows['Building']}"),
            ("EquipmentCode", f"=Config!$B${conf_rows['EquipmentCode']}"),
        ],
        start=2,
    ):
        md.cell(i, 1, key)
        md.cell(i, 2, value)
    md.cell(5, 2).number_format = "DD/MM/YYYY"
    md.sheet_view.showGridLines = False

    for hidden in ("Config", "Lists", "Library"):
        wb[hidden].sheet_state = "hidden"

    # Metadata is machine-readable, not something a reader should be handed. It
    # carries no print area, so LibreOffice printed it -- which put two pages of
    # key/value pairs in front of the cover of every issued PDF. Hidden sheets
    # are still readable by openpyxl and by Power Query, so nothing that consumes
    # it loses anything; a working copy keeps the v1 sheet layout exactly.
    if issue:
        wb["Metadata"].sheet_state = "hidden"
    wb._sheets = [
        wb["Metadata"], wb["Front Cover"], wb["Revision page"], wb["Schedule"],
        wb["Config"], wb["Lists"], wb["Library"],
    ]
    wb.active = 3

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _column_values(content: ScheduleContent, col) -> Iterable[Any]:
    """Every value that will land in this column, for width measurement.

    Includes the computed values as well as the typed ones. An issued export
    writes the snapshot literally and embeds no library, so measuring only the
    typed columns left every product and every calculated column sized from its
    header alone -- which is exactly where the clipping was.
    """
    key, name = col.legacy_name, col.name
    if col.kind == "input":
        for row in content.rows:
            value = row.get(key, row.get(name))
            if value not in (None, ""):
                yield value
    elif col.kind == "library":
        for product in content.products:
            value = product.get(key, product.get(name))
            if value not in (None, ""):
                yield value

    # What each row will actually show, from a snapshot or from an override.
    for computed in content.computed:
        value = computed.get(key, computed.get(name))
        if value not in (None, ""):
            yield _rounded(value)
    for overrides in content.overrides:
        value = overrides.get(key)
        if value not in (None, ""):
            yield value

    if col.unit:
        yield col.unit


def _rounded(value: Any) -> Any:
    """A number as the sheet will show it: two decimals, not seventeen."""
    if isinstance(value, float):
        return f"{value:.2f}"
    return value


def _apply_branding(cover, content: ScheduleContent) -> None:
    """Put the organisation's logo on the cover.

    openpyxl cannot write shapes, but it *can* write images -- so a logo is the
    one piece of branding that survives generation faithfully. The house cover's
    drawn elements still need the cover-template route, and the fonts, colours
    and which fields appear are carried out by the renderer itself rather than
    by anything pasted in here.
    """
    branding = content.branding
    if not branding.logo:
        return

    import base64
    import io

    try:
        from openpyxl.drawing.image import Image as XlImage
    except ImportError:  # Pillow missing
        return

    try:
        logo = branding.logo
        raw = logo.split(",", 1)[1] if logo.startswith("data:") else logo
        stream = io.BytesIO(base64.b64decode(raw))
        image = XlImage(stream)
    except Exception:
        # A broken logo must not stop a schedule being issued.
        return

    scale = branding.logo_scale
    if scale and scale != 1.0:
        image.width = int(image.width * scale)
        image.height = int(image.height * scale)
    cover.add_image(image, branding.logo_anchor or "A1")


def _pretty(unit: str) -> str:
    from ..core.units import pretty_unit

    return pretty_unit(unit)
